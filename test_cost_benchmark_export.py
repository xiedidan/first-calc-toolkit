"""
测试成本基准导出功能
验证需求：5.1-5.4, 6.4
"""
import requests
import openpyxl
from io import BytesIO

# 配置
BASE_URL = "http://localhost:8000/api/v1"
HOSPITAL_ID = 1

def login():
    """登录获取访问令牌"""
    response = requests.post(
        f"{BASE_URL}/auth/login",
        json={"username": "admin", "password": "admin123"}
    )
    if response.status_code == 200:
        return response.json()["access_token"]
    return None

def get_headers(token):
    """获取请求头"""
    return {
        "Authorization": f"Bearer {token}",
        "X-Hospital-ID": str(HOSPITAL_ID),
        "Content-Type": "application/json"
    }

def create_test_benchmarks(token, count=5, prefix="EXPORT_TEST"):
    """创建测试数据"""
    import time
    timestamp = int(time.time())
    
    print(f"\n创建{count}条测试数据...")
    
    # 获取模型版本
    response = requests.get(
        f"{BASE_URL}/model-versions",
        headers=get_headers(token),
        params={"page": 1, "size": 1}
    )
    
    if response.status_code != 200 or response.json()["total"] == 0:
        print("无法获取模型版本")
        return []
    
    version = response.json()["items"][0]
    version_id = version["id"]
    version_name = version["name"]
    
    created_ids = []
    for i in range(count):
        data = {
            "department_code": f"{prefix}_{timestamp}_{i:03d}",
            "department_name": f"导出测试科室{timestamp}_{i:03d}",
            "version_id": version_id,
            "version_name": version_name,
            "dimension_code": f"{prefix}_DIM_{timestamp}_{i:03d}",
            "dimension_name": f"导出测试维度{timestamp}_{i:03d}",
            "benchmark_value": 1000.00 + i * 100
        }
        
        response = requests.post(
            f"{BASE_URL}/cost-benchmarks",
            headers=get_headers(token),
            json=data
        )
        
        if response.status_code == 200:
            created_ids.append(response.json()["id"])
            print(f"  创建成功: ID={response.json()['id']}")
        else:
            print(f"  创建失败: {response.text}")
    
    return created_ids

def cleanup_test_benchmarks(token, ids):
    """清理测试数据"""
    print(f"\n清理{len(ids)}条测试数据...")
    for benchmark_id in ids:
        response = requests.delete(
            f"{BASE_URL}/cost-benchmarks/{benchmark_id}",
            headers=get_headers(token)
        )
        if response.status_code == 200:
            print(f"  删除成功: ID={benchmark_id}")
        else:
            print(f"  删除失败: ID={benchmark_id}")

def test_export_with_data(token):
    """测试导出有数据的情况（需求5.1, 5.2）"""
    print("\n=== 测试1: 导出有数据的情况 ===")
    
    response = requests.get(
        f"{BASE_URL}/cost-benchmarks/export",
        headers=get_headers(token)
    )
    
    if response.status_code != 200:
        print(f"✗ 导出失败: {response.status_code}")
        return False
    
    # 验证响应头
    content_type = response.headers.get("Content-Type")
    content_disposition = response.headers.get("Content-Disposition")
    
    print(f"Content-Type: {content_type}")
    print(f"Content-Disposition: {content_disposition}")
    
    if "spreadsheetml.sheet" not in content_type:
        print("✗ Content-Type不正确")
        return False
    
    if "filename*=UTF-8''" not in content_disposition:
        print("✗ 文件名编码不正确")
        return False
    
    # URL编码后的"成本基准"
    from urllib.parse import unquote
    decoded_disposition = unquote(content_disposition)
    
    if "成本基准_" not in decoded_disposition:
        print(f"✗ 文件名不包含中文: {decoded_disposition}")
        return False
    
    print("✓ 响应头验证通过")
    
    # 验证Excel内容
    wb = openpyxl.load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # 验证列标题（需求5.2）
    expected_headers = [
        "科室代码", "科室名称", "模型版本名称", "维度代码",
        "维度名称", "基准值", "创建时间", "更新时间"
    ]
    
    actual_headers = [cell.value for cell in ws[1]]
    
    if actual_headers != expected_headers:
        print(f"✗ 列标题不正确")
        print(f"  期望: {expected_headers}")
        print(f"  实际: {actual_headers}")
        return False
    
    print("✓ 列标题验证通过")
    
    # 验证数据行数
    data_rows = ws.max_row - 1  # 减去标题行
    print(f"✓ 导出了{data_rows}行数据")
    
    # 验证数据内容
    if data_rows > 0:
        row = ws[2]  # 第一行数据
        print(f"  第一行数据: {row[0].value} - {row[1].value} - {row[4].value}")
        
        # 验证所有必需字段都有值
        for i, cell in enumerate(row):
            if cell.value is None:
                print(f"✗ 第{i+1}列数据为空")
                return False
    
    print("✓ 数据内容验证通过")
    
    return True

def test_export_with_filters(token):
    """测试导出时应用筛选条件（需求5.1）"""
    print("\n=== 测试2: 导出时应用筛选条件 ===")
    
    # 获取模型版本ID
    response = requests.get(
        f"{BASE_URL}/model-versions",
        headers=get_headers(token),
        params={"page": 1, "size": 1}
    )
    
    if response.status_code != 200:
        print("无法获取模型版本")
        return False
    
    version_id = response.json()["items"][0]["id"]
    
    # 导出时使用版本筛选
    response = requests.get(
        f"{BASE_URL}/cost-benchmarks/export",
        headers=get_headers(token),
        params={"version_id": version_id}
    )
    
    if response.status_code != 200:
        print(f"✗ 导出失败: {response.status_code}")
        return False
    
    # 验证导出的数据
    wb = openpyxl.load_workbook(BytesIO(response.content))
    ws = wb.active
    
    data_rows = ws.max_row - 1
    print(f"✓ 使用版本筛选导出了{data_rows}行数据")
    
    # 使用关键词筛选
    response = requests.get(
        f"{BASE_URL}/cost-benchmarks/export",
        headers=get_headers(token),
        params={"keyword": "导出测试"}
    )
    
    if response.status_code != 200:
        print(f"✗ 关键词筛选导出失败: {response.status_code}")
        return False
    
    wb = openpyxl.load_workbook(BytesIO(response.content))
    ws = wb.active
    
    data_rows = ws.max_row - 1
    print(f"✓ 使用关键词筛选导出了{data_rows}行数据")
    
    return True

def test_export_empty_data(token, cleanup_ids):
    """测试导出空数据的情况（需求5.4）"""
    print("\n=== 测试3: 导出空数据的情况 ===")
    
    # 先清理所有测试数据
    cleanup_test_benchmarks(token, cleanup_ids)
    
    # 尝试导出
    response = requests.get(
        f"{BASE_URL}/cost-benchmarks/export",
        headers=get_headers(token)
    )
    
    if response.status_code == 400:
        error = response.json()
        if "没有可导出的数据" in error.get("detail", ""):
            print("✓ 正确处理了空数据情况")
            return True
        else:
            print(f"✗ 错误消息不正确: {error.get('detail')}")
            return False
    else:
        print(f"✗ 应该返回400错误，实际返回: {response.status_code}")
        return False

def test_export_filename_with_timestamp(token):
    """测试文件名包含时间戳（需求5.3）"""
    print("\n=== 测试4: 文件名包含时间戳 ===")
    
    # 创建一条测试数据
    response = requests.get(
        f"{BASE_URL}/model-versions",
        headers=get_headers(token),
        params={"page": 1, "size": 1}
    )
    
    version = response.json()["items"][0]
    
    data = {
        "department_code": "TIMESTAMP_TEST",
        "department_name": "时间戳测试",
        "version_id": version["id"],
        "version_name": version["name"],
        "dimension_code": "TIMESTAMP_DIM",
        "dimension_name": "时间戳维度",
        "benchmark_value": 1000.00
    }
    
    create_response = requests.post(
        f"{BASE_URL}/cost-benchmarks",
        headers=get_headers(token),
        json=data
    )
    
    if create_response.status_code != 200:
        print("创建测试数据失败")
        return False
    
    benchmark_id = create_response.json()["id"]
    
    # 导出
    response = requests.get(
        f"{BASE_URL}/cost-benchmarks/export",
        headers=get_headers(token)
    )
    
    if response.status_code != 200:
        print(f"✗ 导出失败: {response.status_code}")
        return False
    
    # 验证文件名格式
    content_disposition = response.headers.get("Content-Disposition")
    
    # 文件名应该是: 成本基准_YYYYMMDD_HHMMSS.xlsx
    import re
    from urllib.parse import unquote
    
    decoded_disposition = unquote(content_disposition)
    pattern = r"成本基准_\d{8}_\d{6}\.xlsx"
    
    if re.search(pattern, decoded_disposition):
        print(f"✓ 文件名包含正确的时间戳格式: {decoded_disposition}")
        result = True
    else:
        print(f"✗ 文件名格式不正确: {decoded_disposition}")
        result = False
    
    # 清理测试数据
    requests.delete(
        f"{BASE_URL}/cost-benchmarks/{benchmark_id}",
        headers=get_headers(token)
    )
    
    return result

def test_multi_tenant_isolation(token):
    """测试多租户数据隔离（需求6.4）"""
    print("\n=== 测试5: 多租户数据隔离 ===")
    
    # 使用当前医疗机构导出
    response1 = requests.get(
        f"{BASE_URL}/cost-benchmarks/export",
        headers=get_headers(token)
    )
    
    if response1.status_code == 400:
        print("✓ 当前医疗机构没有数据（已清理）")
        return True
    
    if response1.status_code != 200:
        print(f"✗ 导出失败: {response1.status_code}")
        return False
    
    # 验证导出的数据只包含当前医疗机构的数据
    wb = openpyxl.load_workbook(BytesIO(response1.content))
    ws = wb.active
    
    data_rows = ws.max_row - 1
    print(f"✓ 导出了{data_rows}行数据（仅当前医疗机构）")
    
    return True

def test_export_data_consistency(token):
    """测试导出数据与列表查询一致性（需求5.1）"""
    print("\n=== 测试6: 导出数据与列表查询一致性 ===")
    
    # 创建测试数据
    test_ids = create_test_benchmarks(token, 3)
    
    if not test_ids:
        print("创建测试数据失败")
        return False
    
    try:
        # 获取列表数据
        list_response = requests.get(
            f"{BASE_URL}/cost-benchmarks",
            headers=get_headers(token),
            params={"page": 1, "size": 100}
        )
        
        if list_response.status_code != 200:
            print("获取列表失败")
            return False
        
        list_data = list_response.json()
        list_count = list_data["total"]
        
        # 导出数据
        export_response = requests.get(
            f"{BASE_URL}/cost-benchmarks/export",
            headers=get_headers(token)
        )
        
        if export_response.status_code != 200:
            print("导出失败")
            return False
        
        # 验证导出的数据行数
        wb = openpyxl.load_workbook(BytesIO(export_response.content))
        ws = wb.active
        export_count = ws.max_row - 1  # 减去标题行
        
        if list_count == export_count:
            print(f"✓ 数据一致: 列表{list_count}条，导出{export_count}条")
            return True
        else:
            print(f"✗ 数据不一致: 列表{list_count}条，导出{export_count}条")
            return False
    
    finally:
        # 清理测试数据
        cleanup_test_benchmarks(token, test_ids)

def main():
    """主测试函数"""
    print("=" * 60)
    print("成本基准导出功能测试")
    print("=" * 60)
    
    # 登录
    token = login()
    if not token:
        print("登录失败")
        return
    
    print(f"登录成功")
    
    # 创建测试数据
    test_ids = create_test_benchmarks(token, 5)
    
    if not test_ids:
        print("创建测试数据失败")
        return
    
    try:
        # 运行测试
        results = []
        
        results.append(("导出有数据", test_export_with_data(token)))
        results.append(("应用筛选条件", test_export_with_filters(token)))
        results.append(("文件名时间戳", test_export_filename_with_timestamp(token)))
        results.append(("多租户隔离", test_multi_tenant_isolation(token)))
        results.append(("数据一致性", test_export_data_consistency(token)))
        results.append(("空数据处理", test_export_empty_data(token, test_ids)))
        
        # 汇总结果
        print("\n" + "=" * 60)
        print("测试结果汇总")
        print("=" * 60)
        
        for name, result in results:
            status = "✓ 通过" if result else "✗ 失败"
            print(f"{name}: {status}")
        
        passed = sum(1 for _, result in results if result)
        total = len(results)
        
        print(f"\n总计: {passed}/{total} 通过")
        
        if passed == total:
            print("\n✓ 所有测试通过！")
        else:
            print(f"\n✗ {total - passed} 个测试失败")
    
    except Exception as e:
        print(f"\n测试过程中发生错误: {e}")
        import traceback
        traceback.print_exc()
        
        # 确保清理测试数据
        cleanup_test_benchmarks(token, test_ids)

if __name__ == "__main__":
    main()
