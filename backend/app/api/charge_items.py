"""
收费项目管理API
"""
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import or_, desc, asc
import openpyxl

from app.api import deps
from app.models.charge_item import ChargeItem
from app.models.dimension_item_mapping import DimensionItemMapping
from app.schemas.dimension_item import (
    ChargeItem as ChargeItemSchema,
    ChargeItemCreate,
    ChargeItemUpdate,
    ChargeItemList,
)

router = APIRouter()


@router.get("", response_model=ChargeItemList)
def get_charge_items(
    db: Session = Depends(deps.get_db),
    current_user = Depends(deps.get_current_user),
    page: int = Query(1, ge=1, description="页码"),
    size: int = Query(10, ge=1, le=10000, description="每页数量"),
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    item_category: Optional[str] = Query(None, description="项目分类筛选"),
    sort_by: Optional[str] = Query("item_code", description="排序字段"),
    sort_order: Optional[str] = Query("asc", description="排序方向"),
):
    """获取收费项目列表"""
    from app.utils.hospital_filter import apply_hospital_filter
    
    query = db.query(ChargeItem)
    query = apply_hospital_filter(query, ChargeItem, current_user)
    
    # 关键词搜索
    if keyword:
        query = query.filter(
            or_(
                ChargeItem.item_code.contains(keyword),
                ChargeItem.item_name.contains(keyword),
                ChargeItem.item_category.contains(keyword),
            )
        )
    
    # 分类筛选
    if item_category:
        query = query.filter(ChargeItem.item_category == item_category)
    
    # 排序
    sort_column = getattr(ChargeItem, sort_by, ChargeItem.item_code)
    if sort_order == "desc":
        query = query.order_by(desc(sort_column))
    else:
        query = query.order_by(asc(sort_column))
    
    # 先获取分页数据（这个很快）
    items = query.offset((page - 1) * size).limit(size).all()
    
    # 总数查询（这个可能慢）
    # 优化：使用子查询避免重复的 JOIN 和过滤
    total = query.with_entities(ChargeItem.id).count()
    
    return ChargeItemList(total=total, items=items)


@router.post("", response_model=ChargeItemSchema)
def create_charge_item(
    item_in: ChargeItemCreate,
    db: Session = Depends(deps.get_db),
    current_user = Depends(deps.get_current_user),
):
    """创建收费项目"""
    from app.utils.hospital_filter import get_user_hospital_id
    
    hospital_id = get_user_hospital_id(current_user)
    
    # 检查项目编码是否已存在（同一医疗机构内）
    existing = db.query(ChargeItem).filter(
        ChargeItem.hospital_id == hospital_id,
        ChargeItem.item_code == item_in.item_code
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="收费项目编码已存在")
    
    # 创建收费项目
    item_data = item_in.model_dump()
    item_data['hospital_id'] = hospital_id
    item = ChargeItem(**item_data)
    db.add(item)
    db.commit()
    db.refresh(item)
    
    return item


@router.delete("/clear-all")
def clear_all_charge_items(
    db: Session = Depends(deps.get_db),
    current_user = Depends(deps.get_current_user),
):
    """清空所有收费项目（仅用于测试）"""
    from app.utils.hospital_filter import get_user_hospital_id
    
    try:
        hospital_id = get_user_hospital_id(current_user)
        
        # 先删除当前医疗机构的维度目录映射
        charge_item_ids = db.query(ChargeItem.id).filter(
            ChargeItem.hospital_id == hospital_id
        ).all()
        charge_item_ids = [item[0] for item in charge_item_ids]
        
        if charge_item_ids:
            db.query(DimensionItemMapping).filter(
                DimensionItemMapping.charge_item_id.in_(charge_item_ids)
            ).delete(synchronize_session=False)
        
        # 再删除当前医疗机构的收费项目
        count = db.query(ChargeItem).filter(
            ChargeItem.hospital_id == hospital_id
        ).delete()
        
        db.commit()
        
        return {
            "message": f"成功清空 {count} 条收费项目数据",
            "deleted_count": count
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"清空失败: {str(e)}")


@router.get("/{item_id}", response_model=ChargeItemSchema)
def get_charge_item(
    item_id: int,
    db: Session = Depends(deps.get_db),
    current_user = Depends(deps.get_current_user),
):
    """获取收费项目详情"""
    from app.utils.hospital_filter import get_user_hospital_id
    
    hospital_id = get_user_hospital_id(current_user)
    item = db.query(ChargeItem).filter(
        ChargeItem.id == item_id,
        ChargeItem.hospital_id == hospital_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="收费项目不存在")
    
    return item


@router.put("/{item_id}", response_model=ChargeItemSchema)
def update_charge_item(
    item_id: int,
    item_in: ChargeItemUpdate,
    db: Session = Depends(deps.get_db),
    current_user = Depends(deps.get_current_user),
):
    """更新收费项目信息"""
    from app.utils.hospital_filter import get_user_hospital_id
    
    hospital_id = get_user_hospital_id(current_user)
    item = db.query(ChargeItem).filter(
        ChargeItem.id == item_id,
        ChargeItem.hospital_id == hospital_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="收费项目不存在")
    
    # 更新字段
    update_data = item_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(item, field, value)
    
    db.commit()
    db.refresh(item)
    
    return item


@router.delete("/{item_id}")
def delete_charge_item(
    item_id: int,
    db: Session = Depends(deps.get_db),
    current_user = Depends(deps.get_current_user),
):
    """删除收费项目"""
    from app.utils.hospital_filter import get_user_hospital_id
    
    hospital_id = get_user_hospital_id(current_user)
    item = db.query(ChargeItem).filter(
        ChargeItem.id == item_id,
        ChargeItem.hospital_id == hospital_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="收费项目不存在")
    
    # 检查是否被维度目录引用
    mapping_count = db.query(DimensionItemMapping).filter(
        DimensionItemMapping.charge_item_id == item.id
    ).count()
    if mapping_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"该收费项目已被 {mapping_count} 个维度目录引用，无法删除"
        )
    
    db.delete(item)
    db.commit()
    
    return {"message": "删除成功"}


@router.get("/categories/list", response_model=list[str])
def get_categories(
    db: Session = Depends(deps.get_db),
    current_user = Depends(deps.get_current_user),
):
    """获取所有分类列表"""
    from app.utils.hospital_filter import get_user_hospital_id
    
    hospital_id = get_user_hospital_id(current_user)
    categories = db.query(ChargeItem.item_category).distinct().filter(
        ChargeItem.hospital_id == hospital_id,
        ChargeItem.item_category.isnot(None)
    ).all()
    return [cat[0] for cat in categories if cat[0]]


@router.post("/parse")
async def parse_excel(
    file: UploadFile = File(...),
    db: Session = Depends(deps.get_db),
    current_user = Depends(deps.get_current_user),
):
    """解析Excel文件"""
    from app.services.excel_import_service import ExcelImportService
    from app.config.import_configs import CHARGE_ITEM_IMPORT_CONFIG
    
    # 验证文件格式
    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(
            status_code=400, 
            detail="仅支持 .xlsx 格式文件，请使用 Excel 2007 及以上版本保存"
        )
    
    # 读取文件内容
    content = await file.read()
    
    # 创建导入服务
    service = ExcelImportService(CHARGE_ITEM_IMPORT_CONFIG)
    
    try:
        result = service.parse_excel(content)
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析Excel失败: {str(e)}")


@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    mapping: str = Form(...),
    async_mode: bool = Form(True),
    db: Session = Depends(deps.get_db),
    current_user = Depends(deps.get_current_user),
):
    """批量导入收费项目（支持同步和异步模式）"""
    import json
    from app.services.excel_import_service import ExcelImportService
    from app.config.import_configs import CHARGE_ITEM_IMPORT_CONFIG
    from app.utils.hospital_filter import get_user_hospital_id
    
    hospital_id = get_user_hospital_id(current_user)
    
    # 验证文件格式
    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(
            status_code=400, 
            detail="仅支持 .xlsx 格式文件，请使用 Excel 2007 及以上版本保存"
        )
    
    # 解析映射关系
    try:
        mapping_dict = json.loads(mapping)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="映射关系格式错误")
    
    # 读取文件内容
    content = await file.read()
    
    # 创建导入服务
    service = ExcelImportService(CHARGE_ITEM_IMPORT_CONFIG)
    
    # 验证映射
    validation = service.validate_mapping(mapping_dict)
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail="; ".join(validation["errors"]))
    
    # 异步模式：提交到 Celery 任务队列
    if async_mode:
        from app.tasks.import_tasks import import_charge_items_task
        
        # 提交异步任务（传递hospital_id）
        task = import_charge_items_task.delay(content, mapping_dict, hospital_id)
        
        return {
            "task_id": task.id,
            "status": "pending",
            "message": "导入任务已提交，请使用 task_id 查询进度"
        }
    
    # 同步模式：直接执行导入（用于小数据量）
    else:
        # 自定义验证函数
        def validate_charge_item(row_data: dict, db: Session) -> Optional[str]:
            item_code = row_data.get("item_code", "")
            if item_code:
                existing = db.query(ChargeItem).filter(
                    ChargeItem.hospital_id == hospital_id,
                    ChargeItem.item_code == item_code
                ).first()
                if existing:
                    return f"项目编码 {item_code} 已存在"
            return None
        
        # 数据预处理函数：添加hospital_id
        def preprocess_row(row_data: dict) -> dict:
            row_data['hospital_id'] = hospital_id
            return row_data
        
        try:
            result = service.import_data(
                content,
                mapping_dict,
                db,
                ChargeItem,
                validate_charge_item,
                preprocess_row
            )
            return result
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"导入失败: {str(e)}")


@router.get("/import/status/{task_id}")
def get_import_status(
    task_id: str,
    current_user = Depends(deps.get_current_user),
):
    """查询导入任务状态"""
    from app.celery_app import celery_app
    
    try:
        # 使用配置好的 celery_app 实例
        task = celery_app.AsyncResult(task_id)
        
        # 检查 backend 是否可用
        backend_type = type(task.backend).__name__
        if backend_type == 'DisabledBackend':
            raise HTTPException(
                status_code=503,
                detail="Celery result backend 未启用。请检查：1) Redis 是否运行 2) .env 中 CELERY_RESULT_BACKEND 配置 3) 重启 FastAPI 服务"
            )
        
        if task.state == 'PENDING':
            response = {
                'state': task.state,
                'status': '任务等待中...',
                'current': 0,
                'total': 0
            }
        elif task.state == 'PROCESSING':
            response = {
                'state': task.state,
                'status': task.info.get('status', ''),
                'current': task.info.get('current', 0),
                'total': task.info.get('total', 0)
            }
        elif task.state == 'SUCCESS':
            response = {
                'state': task.state,
                'status': '导入完成',
                'result': task.info
            }
        elif task.state == 'FAILURE':
            response = {
                'state': task.state,
                'status': '导入失败',
                'error': str(task.info)
            }
        else:
            response = {
                'state': task.state,
                'status': str(task.info)
            }
        
        return response
        
    except AttributeError as e:
        raise HTTPException(
            status_code=503,
            detail=f"Celery backend 配置错误: {str(e)}。请确保 Redis 正在运行并且 CELERY_RESULT_BACKEND 已正确配置。"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"查询任务状态失败: {str(e)}"
        )


@router.get("/template")
def download_template(
    current_user = Depends(deps.get_current_user),
):
    """下载导入模板"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from io import BytesIO
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收费项目"
    
    # 设置表头
    headers = ["项目编码", "项目名称", "项目分类", "单价"]
    ws.append(headers)
    
    # 添加示例数据
    ws.append(["CK001", "血常规", "检验", "25.00"])
    ws.append(["CK002", "尿常规", "检验", "15.00"])
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    
    # 添加批注说明
    ws['A1'].comment = openpyxl.comments.Comment("必填，唯一", "系统")
    ws['B1'].comment = openpyxl.comments.Comment("必填", "系统")
    ws['C1'].comment = openpyxl.comments.Comment("可选", "系统")
    ws['D1'].comment = openpyxl.comments.Comment("可选", "系统")
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=charge_items_template.xlsx"}
    )
