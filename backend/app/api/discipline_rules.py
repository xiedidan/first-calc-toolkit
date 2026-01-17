"""
学科规则API
"""
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.discipline_rule import DisciplineRule
from app.models.model_version import ModelVersion
from app.models.user import User
from app.schemas.discipline_rule import (
    DisciplineRuleCreate,
    DisciplineRuleUpdate,
    DisciplineRuleResponse,
    DisciplineRuleListResponse,
)
from app.api.deps import get_current_user
from app.utils.hospital_filter import (
    apply_hospital_filter,
    get_current_hospital_id_or_raise,
    validate_hospital_access,
)

router = APIRouter()


def _batch_get_dimension_full_paths(db: Session, dimension_codes: list[str]) -> dict[str, str]:
    """批量获取维度节点的完整层级路径（优化版本，减少数据库查询）
    
    返回: {dimension_code: "序列 - 一级维度 - 二级维度..."}
    """
    if not dimension_codes:
        return {}
    
    from app.models.model_node import ModelNode
    
    # 1. 查询所有目标节点
    target_nodes = db.query(ModelNode).filter(ModelNode.code.in_(dimension_codes)).all()
    if not target_nodes:
        return {}
    
    # 2. 收集所有需要的节点ID（包括所有祖先）
    # 先获取目标节点所属的版本ID
    version_ids = set(node.version_id for node in target_nodes)
    
    # 3. 一次性查询这些版本的所有节点，构建内存中的节点映射
    all_nodes = db.query(ModelNode).filter(ModelNode.version_id.in_(version_ids)).all()
    nodes_by_id = {node.id: node for node in all_nodes}
    
    # 4. 为每个目标节点构建完整路径
    result = {}
    for node in target_nodes:
        path_parts = []
        node_types = []
        current = node
        
        # 在内存中向上遍历
        while current:
            path_parts.append(current.name)
            node_types.append(current.node_type)
            if current.parent_id and current.parent_id in nodes_by_id:
                current = nodes_by_id[current.parent_id]
            else:
                break
        
        # 反转路径和类型
        path_parts.reverse()
        node_types.reverse()
        
        # 找到第一个序列节点的位置
        start_idx = 0
        for i, nt in enumerate(node_types):
            if nt == 'sequence':
                start_idx = i
                break
        
        # 构建路径
        if start_idx < len(path_parts):
            result[node.code] = " - ".join(path_parts[start_idx:])
        else:
            result[node.code] = path_parts[-1] if path_parts else ""
    
    return result


@router.delete("/batch-delete")
def batch_delete_discipline_rules(
    version_id: Optional[int] = Query(None, description="模型版本ID"),
    department_code: Optional[str] = Query(None, description="科室代码"),
    dimension_code: Optional[str] = Query(None, description="维度代码"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """批量删除筛选出的学科规则"""
    query = db.query(DisciplineRule)
    query = apply_hospital_filter(query, DisciplineRule, required=True)
    
    # 筛选条件（与列表查询一致）
    if version_id:
        query = query.filter(DisciplineRule.version_id == version_id)
    if department_code:
        query = query.filter(DisciplineRule.department_code == department_code)
    if dimension_code:
        query = query.filter(DisciplineRule.dimension_code == dimension_code)
    if keyword:
        keyword_filter = f"%{keyword}%"
        query = query.filter(
            (DisciplineRule.department_name.ilike(keyword_filter)) |
            (DisciplineRule.dimension_name.ilike(keyword_filter)) |
            (DisciplineRule.rule_description.ilike(keyword_filter))
        )
    
    # 统计要删除的数量
    count = query.count()
    
    if count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="没有符合条件的学科规则"
        )
    
    # 执行删除
    query.delete(synchronize_session=False)
    db.commit()
    
    return {"deleted_count": count}


@router.get("", response_model=DisciplineRuleListResponse)
def get_discipline_rules(
    page: int = Query(1, ge=1, description="页码"),
    size: int = Query(20, ge=1, le=1000, description="每页数量"),
    version_id: Optional[int] = Query(None, description="模型版本ID"),
    department_code: Optional[str] = Query(None, description="科室代码"),
    dimension_code: Optional[str] = Query(None, description="维度代码"),
    node_id: Optional[int] = Query(None, description="节点ID（用于查询维度代码）"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取学科规则列表"""
    query = db.query(DisciplineRule)
    query = apply_hospital_filter(query, DisciplineRule, required=True)
    
    # 如果传入了 node_id，先查询对应的 dimension_code
    if node_id and not dimension_code:
        from app.models.model_node import ModelNode
        node = db.query(ModelNode).filter(ModelNode.id == node_id).first()
        if node:
            dimension_code = node.code
    
    # 筛选条件
    if version_id:
        query = query.filter(DisciplineRule.version_id == version_id)
    if department_code:
        query = query.filter(DisciplineRule.department_code == department_code)
    if dimension_code:
        query = query.filter(DisciplineRule.dimension_code == dimension_code)
    if keyword:
        keyword_filter = f"%{keyword}%"
        query = query.filter(
            (DisciplineRule.department_name.ilike(keyword_filter)) |
            (DisciplineRule.dimension_name.ilike(keyword_filter)) |
            (DisciplineRule.rule_description.ilike(keyword_filter))
        )
    
    # 统计总数
    total = query.count()
    
    # 分页
    query = query.order_by(DisciplineRule.id.desc())
    items = query.offset((page - 1) * size).limit(size).all()
    
    # 预加载版本名称
    version_ids = list(set(item.version_id for item in items))
    versions = {v.id: v.name for v in db.query(ModelVersion).filter(ModelVersion.id.in_(version_ids)).all()}
    
    # 预加载维度节点，用于计算完整路径（批量优化）
    dimension_codes = list(set(item.dimension_code for item in items))
    dimension_full_paths = _batch_get_dimension_full_paths(db, dimension_codes)
    
    result_items = []
    for item in items:
        # 优先使用动态计算的完整路径，如果没有则使用存储的名称
        dimension_name = dimension_full_paths.get(item.dimension_code, item.dimension_name)
        
        item_dict = {
            "id": item.id,
            "department_code": item.department_code,
            "department_name": item.department_name,
            "version_id": item.version_id,
            "version_name": versions.get(item.version_id, ""),
            "dimension_code": item.dimension_code,
            "dimension_name": dimension_name,
            "rule_description": item.rule_description,
            "rule_coefficient": item.rule_coefficient,
            "created_at": item.created_at,
            "updated_at": item.updated_at,
        }
        result_items.append(item_dict)
    
    return {"items": result_items, "total": total}


@router.get("/export")
def export_discipline_rules(
    version_id: Optional[int] = Query(None, description="模型版本ID"),
    department_code: Optional[str] = Query(None, description="科室代码"),
    dimension_code: Optional[str] = Query(None, description="维度代码"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出学科规则为Excel"""
    from io import BytesIO
    from datetime import datetime
    from urllib.parse import quote
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from fastapi.responses import StreamingResponse
    from app.models.hospital import Hospital
    from app.models.department import Department
    from app.models.dimension_analysis import DimensionAnalysis
    import sqlalchemy as sa
    
    hospital_id = get_current_hospital_id_or_raise()
    
    # 获取医院名称
    hospital = db.query(Hospital).filter(Hospital.id == hospital_id).first()
    hospital_name = hospital.name if hospital else ""
    
    query = db.query(DisciplineRule)
    query = apply_hospital_filter(query, DisciplineRule, required=True)
    
    # 筛选条件
    if version_id:
        query = query.filter(DisciplineRule.version_id == version_id)
    if department_code:
        query = query.filter(DisciplineRule.department_code == department_code)
    if dimension_code:
        query = query.filter(DisciplineRule.dimension_code == dimension_code)
    if keyword:
        keyword_filter = f"%{keyword}%"
        query = query.filter(
            (DisciplineRule.department_name.ilike(keyword_filter)) |
            (DisciplineRule.dimension_name.ilike(keyword_filter)) |
            (DisciplineRule.rule_description.ilike(keyword_filter))
        )
    
    # 按版本和科室排序
    query = query.order_by(DisciplineRule.version_id, DisciplineRule.department_code, DisciplineRule.id)
    items = query.all()
    
    # 预加载版本名称
    version_ids = list(set(item.version_id for item in items))
    versions = {v.id: v.name for v in db.query(ModelVersion).filter(ModelVersion.id.in_(version_ids)).all()}
    
    # 预加载维度节点，用于计算完整路径（批量优化）
    dimension_codes_list = list(set(item.dimension_code for item in items))
    dimension_full_paths = _batch_get_dimension_full_paths(db, dimension_codes_list)
    
    # 预加载业务分析数据
    dept_codes = list(set(item.department_code for item in items))
    
    # 通过 department_code 查找 department_id（同时匹配 his_code 和 accounting_unit_code）
    departments = db.query(Department).filter(
        Department.hospital_id == hospital_id,
        sa.or_(
            Department.his_code.in_(dept_codes),
            Department.accounting_unit_code.in_(dept_codes),
        ),
    ).all()
    
    dept_code_to_id = {}
    for d in departments:
        if d.his_code in dept_codes:
            dept_code_to_id[d.his_code] = d.id
        if d.accounting_unit_code and d.accounting_unit_code in dept_codes:
            dept_code_to_id[d.accounting_unit_code] = d.id
    
    # 通过 dimension_code 查找 node_id
    from app.models.model_node import ModelNode
    nodes = db.query(ModelNode).filter(ModelNode.code.in_(dimension_codes_list)).all()
    dim_code_to_id = {n.code: n.id for n in nodes}
    
    # 查询所有相关的维度分析（长期分析）
    dept_ids = list(set(dept_code_to_id.values()))
    node_ids = list(dim_code_to_id.values())
    
    analysis_map = {}  # {dept_code|dim_code: content}
    if dept_ids and node_ids:
        analyses = db.query(DimensionAnalysis).filter(
            DimensionAnalysis.hospital_id == hospital_id,
            DimensionAnalysis.department_id.in_(dept_ids),
            DimensionAnalysis.node_id.in_(node_ids),
            DimensionAnalysis.period.is_(None),  # 只取长期分析
        ).all()
        
        # 构建反向映射
        id_to_dept_codes = {}
        for code, dept_id in dept_code_to_id.items():
            if dept_id not in id_to_dept_codes:
                id_to_dept_codes[dept_id] = []
            id_to_dept_codes[dept_id].append(code)
        id_to_dim_code = {v: k for k, v in dim_code_to_id.items()}
        
        for analysis in analyses:
            dept_codes_for_id = id_to_dept_codes.get(analysis.department_id, [])
            dim_code = id_to_dim_code.get(analysis.node_id)
            if dept_codes_for_id and dim_code:
                for dc in dept_codes_for_id:
                    key = f"{dc}|{dim_code}"
                    analysis_map[key] = analysis.content
    
    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "学科规则"
    
    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头（与页面列一致）
    headers = ["模型版本", "科室名称", "维度名称", "规则描述", "规则参数", "业务分析", "更新时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 数据行
    for row_idx, item in enumerate(items, 2):
        dimension_name = dimension_full_paths.get(item.dimension_code, item.dimension_name)
        analysis_key = f"{item.department_code}|{item.dimension_code}"
        analysis_content = analysis_map.get(analysis_key, "")
        
        # 格式化更新时间
        updated_at_str = ""
        if item.updated_at:
            updated_at_str = item.updated_at.strftime("%Y-%m-%d %H:%M:%S")
        
        row_data = [
            versions.get(item.version_id, ""),
            item.department_name,
            dimension_name,
            item.rule_description or "",
            float(item.rule_coefficient),
            analysis_content,
            updated_at_str,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if col == 5:  # 规则参数列
                cell.number_format = '0.0000'
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20  # 模型版本
    ws.column_dimensions['B'].width = 18  # 科室名称
    ws.column_dimensions['C'].width = 45  # 维度名称
    ws.column_dimensions['D'].width = 30  # 规则描述
    ws.column_dimensions['E'].width = 12  # 规则参数
    ws.column_dimensions['F'].width = 40  # 业务分析
    ws.column_dimensions['G'].width = 20  # 更新时间
    
    # 保存到内存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 生成文件名：{医院名称}_学科规则_{版本名称}.xlsx
    version_name = versions.get(version_id, "") if version_id else ""
    if hospital_name and version_name:
        filename = f"{hospital_name}_学科规则_{version_name}.xlsx"
    elif hospital_name:
        filename = f"{hospital_name}_学科规则.xlsx"
    elif version_name:
        filename = f"学科规则_{version_name}.xlsx"
    else:
        filename = "学科规则.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        }
    )


@router.post("", response_model=DisciplineRuleResponse, status_code=status.HTTP_201_CREATED)
def create_discipline_rule(
    rule_in: DisciplineRuleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """创建学科规则"""
    hospital_id = get_current_hospital_id_or_raise()
    
    # 验证模型版本存在
    version = db.query(ModelVersion).filter(ModelVersion.id == rule_in.version_id).first()
    if not version:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="模型版本不存在"
        )
    validate_hospital_access(db, version, hospital_id)
    
    # 检查是否已存在相同的规则
    existing = db.query(DisciplineRule).filter(
        DisciplineRule.hospital_id == hospital_id,
        DisciplineRule.version_id == rule_in.version_id,
        DisciplineRule.department_code == rule_in.department_code,
        DisciplineRule.dimension_code == rule_in.dimension_code,
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="该科室和维度的学科规则已存在"
        )
    
    # 创建规则
    db_rule = DisciplineRule(
        hospital_id=hospital_id,
        version_id=rule_in.version_id,
        department_code=rule_in.department_code,
        department_name=rule_in.department_name,
        dimension_code=rule_in.dimension_code,
        dimension_name=rule_in.dimension_name,
        rule_description=rule_in.rule_description,
        rule_coefficient=rule_in.rule_coefficient,
    )
    db.add(db_rule)
    db.commit()
    db.refresh(db_rule)
    
    return {
        **db_rule.__dict__,
        "version_name": version.name,
    }


@router.get("/{rule_id}", response_model=DisciplineRuleResponse)
def get_discipline_rule(
    rule_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取学科规则详情"""
    query = db.query(DisciplineRule).filter(DisciplineRule.id == rule_id)
    query = apply_hospital_filter(query, DisciplineRule, required=True)
    rule = query.first()
    
    if not rule:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="学科规则不存在"
        )
    
    version = db.query(ModelVersion).filter(ModelVersion.id == rule.version_id).first()
    
    return {
        **rule.__dict__,
        "version_name": version.name if version else "",
    }


@router.put("/{rule_id}", response_model=DisciplineRuleResponse)
def update_discipline_rule(
    rule_id: int,
    rule_in: DisciplineRuleUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """更新学科规则"""
    query = db.query(DisciplineRule).filter(DisciplineRule.id == rule_id)
    query = apply_hospital_filter(query, DisciplineRule, required=True)
    rule = query.first()
    
    if not rule:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="学科规则不存在"
        )
    
    validate_hospital_access(db, rule)
    
    # 更新字段
    update_data = rule_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(rule, field, value)
    
    db.commit()
    db.refresh(rule)
    
    version = db.query(ModelVersion).filter(ModelVersion.id == rule.version_id).first()
    
    return {
        **rule.__dict__,
        "version_name": version.name if version else "",
    }


@router.delete("/{rule_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_discipline_rule(
    rule_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """删除学科规则"""
    query = db.query(DisciplineRule).filter(DisciplineRule.id == rule_id)
    query = apply_hospital_filter(query, DisciplineRule, required=True)
    rule = query.first()
    
    if not rule:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="学科规则不存在"
        )
    
    validate_hospital_access(db, rule)
    
    db.delete(rule)
    db.commit()
