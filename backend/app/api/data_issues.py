"""
数据问题记录 API endpoints
"""
from typing import Optional
from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_

from app.api.deps import get_db, get_current_active_user
from app.models import User, DataIssue, ProcessingStage
from app.schemas import (
    DataIssue as DataIssueSchema,
    DataIssueCreate,
    DataIssueUpdate,
    DataIssueList,
)
from app.middleware import get_current_hospital_id

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


router = APIRouter()


@router.get("", response_model=DataIssueList)
async def get_data_issues(
    page: int = Query(1, ge=1, description="页码"),
    size: int = Query(20, ge=1, le=100, description="每页数量"),
    keyword: Optional[str] = Query(None, description="关键词搜索（标题、描述）"),
    processing_stage: Optional[str] = Query(None, description="处理阶段筛选"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    获取数据问题记录列表（分页、筛选）
    """
    # 构建查询
    query = db.query(DataIssue)
    
    # 多租户数据隔离：优先使用激活的医疗机构，否则使用用户所属医疗机构
    hospital_id = get_current_hospital_id()
    if hospital_id is None:
        hospital_id = current_user.hospital_id
    
    # 如果有医疗机构限制，则过滤数据
    if hospital_id is not None:
        query = query.filter(DataIssue.hospital_id == hospital_id)
    
    # 关键词搜索（标题和描述）
    if keyword:
        query = query.filter(
            or_(
                DataIssue.title.ilike(f"%{keyword}%"),
                DataIssue.description.ilike(f"%{keyword}%")
            )
        )
    
    # 处理阶段筛选
    if processing_stage:
        try:
            stage_enum = ProcessingStage(processing_stage)
            query = query.filter(DataIssue.processing_stage == stage_enum)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"无效的处理阶段: {processing_stage}"
            )
    
    # 获取总数
    total = query.count()
    
    # 分页和排序（按创建时间倒序）
    issues = query.order_by(DataIssue.created_at.desc()).offset((page - 1) * size).limit(size).all()
    
    # 格式化响应
    items = []
    for issue in issues:
        items.append(DataIssueSchema(
            id=issue.id,
            title=issue.title,
            description=issue.description,
            reporter=issue.reporter,
            reporter_user_id=issue.reporter_user_id,
            assignee=issue.assignee,
            assignee_user_id=issue.assignee_user_id,
            processing_stage=issue.processing_stage,
            resolution=issue.resolution,
            hospital_id=issue.hospital_id,
            created_at=issue.created_at,
            resolved_at=issue.resolved_at,
            updated_at=issue.updated_at
        ))
    
    return DataIssueList(total=total, items=items)



@router.post("", response_model=DataIssueSchema, status_code=status.HTTP_201_CREATED)
async def create_data_issue(
    issue_create: DataIssueCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    创建新的数据问题记录
    """
    # 获取当前激活的医疗机构ID（从请求头X-Hospital-ID）
    hospital_id = get_current_hospital_id()
    
    # 如果没有激活医疗机构，使用用户所属的医疗机构
    if hospital_id is None:
        if current_user.hospital_id is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="请先激活医疗机构"
            )
        hospital_id = current_user.hospital_id
    
    # 创建问题记录
    issue = DataIssue(
        title=issue_create.title,
        description=issue_create.description,
        reporter=issue_create.reporter,
        reporter_user_id=issue_create.reporter_user_id,
        assignee=issue_create.assignee,
        assignee_user_id=issue_create.assignee_user_id,
        processing_stage=ProcessingStage(issue_create.processing_stage.value) if isinstance(issue_create.processing_stage, str) else issue_create.processing_stage,
        resolution=issue_create.resolution,
        hospital_id=hospital_id
    )
    
    db.add(issue)
    db.commit()
    db.refresh(issue)
    
    return DataIssueSchema(
        id=issue.id,
        title=issue.title,
        description=issue.description,
        reporter=issue.reporter,
        reporter_user_id=issue.reporter_user_id,
        assignee=issue.assignee,
        assignee_user_id=issue.assignee_user_id,
        processing_stage=issue.processing_stage,
        resolution=issue.resolution,
        hospital_id=issue.hospital_id,
        created_at=issue.created_at,
        resolved_at=issue.resolved_at,
        updated_at=issue.updated_at
    )


@router.get("/export")
async def export_data_issues(
    keyword: Optional[str] = None,
    processing_stage: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    导出数据问题记录为Excel文件
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel导出功能不可用，请安装openpyxl库"
        )
    
    # 构建查询（复用列表查询的逻辑）
    query = db.query(DataIssue)
    
    # 多租户数据隔离
    hospital_id = get_current_hospital_id()
    if hospital_id is None:
        hospital_id = current_user.hospital_id
    
    if hospital_id is not None:
        query = query.filter(DataIssue.hospital_id == hospital_id)
    
    # 关键词搜索
    if keyword:
        query = query.filter(
            or_(
                DataIssue.title.ilike(f"%{keyword}%"),
                DataIssue.description.ilike(f"%{keyword}%")
            )
        )
    
    # 处理阶段筛选
    if processing_stage:
        try:
            stage_enum = ProcessingStage(processing_stage)
            query = query.filter(DataIssue.processing_stage == stage_enum)
        except ValueError:
            pass
    
    # 获取所有数据并按创建时间倒序排列
    issues = query.order_by(DataIssue.created_at.desc()).all()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "数据问题记录"
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='微软雅黑', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 设置表头
    headers_list = ['编号', '标题', '问题描述', '记录人', '负责人', '处理阶段', '解决方案', '记录时间', '解决时间']
    for col_num, header in enumerate(headers_list, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    
    # 处理阶段标签映射
    stage_labels = {
        ProcessingStage.NOT_STARTED: '待开始',
        ProcessingStage.IN_PROGRESS: '进行中',
        ProcessingStage.RESOLVED: '已解决',
        ProcessingStage.CONFIRMED: '已确认'
    }
    
    # 填充数据
    for index, issue in enumerate(issues, 1):
        row_num = index + 1
        
        # 转换UTC时间为中国时区 (UTC+8)
        created_at_local = (issue.created_at + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S') if issue.created_at else ''
        resolved_at_local = (issue.resolved_at + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S') if issue.resolved_at else ''
        
        data = [
            index,  # 使用序号而不是数据库ID
            issue.title,
            issue.description,
            issue.reporter,
            issue.assignee or '待定',
            stage_labels.get(issue.processing_stage, issue.processing_stage.value),
            issue.resolution or '',
            created_at_local,
            resolved_at_local
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 获取医院名称
    from app.models.hospital import Hospital
    if hospital_id:
        hospital = db.query(Hospital).filter(Hospital.id == hospital_id).first()
        hospital_name = hospital.name if hospital else "未知医院"
    else:
        hospital_name = "未知医院"
    
    # 返回文件 - 使用URL编码处理中文文件名（医院名称_数据问题记录_日期.xlsx）
    filename = f"{hospital_name}_数据问题记录_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    encoded_filename = quote(filename)
    response_headers = {
        'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"
    }
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=response_headers
    )


@router.get("/{issue_id}", response_model=DataIssueSchema)
async def get_data_issue(
    issue_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    获取数据问题记录详情
    """
    issue = db.query(DataIssue).filter(DataIssue.id == issue_id).first()
    
    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="问题记录不存在"
        )
    
    # 验证访问权限：检查是否有权访问该医疗机构的数据
    hospital_id = get_current_hospital_id()
    if hospital_id is None:
        hospital_id = current_user.hospital_id
    
    if hospital_id is not None and issue.hospital_id != hospital_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权访问该问题记录"
        )
    
    return DataIssueSchema(
        id=issue.id,
        title=issue.title,
        description=issue.description,
        reporter=issue.reporter,
        reporter_user_id=issue.reporter_user_id,
        assignee=issue.assignee,
        assignee_user_id=issue.assignee_user_id,
        processing_stage=issue.processing_stage,
        resolution=issue.resolution,
        hospital_id=issue.hospital_id,
        created_at=issue.created_at,
        resolved_at=issue.resolved_at,
        updated_at=issue.updated_at
    )



@router.put("/{issue_id}", response_model=DataIssueSchema)
async def update_data_issue(
    issue_id: int,
    issue_update: DataIssueUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    更新数据问题记录
    """
    issue = db.query(DataIssue).filter(DataIssue.id == issue_id).first()
    
    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="问题记录不存在"
        )
    
    # 验证访问权限
    hospital_id = get_current_hospital_id()
    if hospital_id is None:
        hospital_id = current_user.hospital_id
    
    if hospital_id is not None and issue.hospital_id != hospital_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权修改该问题记录"
        )
    
    # 记录原始的processing_stage
    old_stage = issue.processing_stage
    
    # 更新字段
    if issue_update.title is not None:
        issue.title = issue_update.title
    
    if issue_update.description is not None:
        issue.description = issue_update.description
    
    if issue_update.reporter is not None:
        issue.reporter = issue_update.reporter
    
    if issue_update.reporter_user_id is not None:
        issue.reporter_user_id = issue_update.reporter_user_id
    
    if issue_update.assignee is not None:
        issue.assignee = issue_update.assignee
    
    if issue_update.assignee_user_id is not None:
        issue.assignee_user_id = issue_update.assignee_user_id
    
    if issue_update.processing_stage is not None:
        issue.processing_stage = issue_update.processing_stage
        
        # 当processing_stage首次更新为"resolved"时，自动设置resolved_at
        if issue_update.processing_stage == ProcessingStage.RESOLVED and old_stage != ProcessingStage.RESOLVED:
            issue.resolved_at = datetime.utcnow()
        # 如果从resolved改为其他状态，清除resolved_at
        elif issue_update.processing_stage != ProcessingStage.RESOLVED and old_stage == ProcessingStage.RESOLVED:
            issue.resolved_at = None
    
    if issue_update.resolution is not None:
        issue.resolution = issue_update.resolution
    
    db.commit()
    db.refresh(issue)
    
    return DataIssueSchema(
        id=issue.id,
        title=issue.title,
        description=issue.description,
        reporter=issue.reporter,
        reporter_user_id=issue.reporter_user_id,
        assignee=issue.assignee,
        assignee_user_id=issue.assignee_user_id,
        processing_stage=issue.processing_stage,
        resolution=issue.resolution,
        hospital_id=issue.hospital_id,
        created_at=issue.created_at,
        resolved_at=issue.resolved_at,
        updated_at=issue.updated_at
    )



@router.delete("/{issue_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_data_issue(
    issue_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    删除数据问题记录
    """
    issue = db.query(DataIssue).filter(DataIssue.id == issue_id).first()
    
    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="问题记录不存在"
        )
    
    # 验证访问权限
    hospital_id = get_current_hospital_id()
    if hospital_id is None:
        hospital_id = current_user.hospital_id
    
    if hospital_id is not None and issue.hospital_id != hospital_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权删除该问题记录"
        )
    
    db.delete(issue)
    db.commit()
    
    return None
