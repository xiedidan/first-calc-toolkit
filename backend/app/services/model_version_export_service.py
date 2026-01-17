"""
模型版本导出服务
"""
from io import BytesIO
from typing import Dict, List, Optional, Set, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from sqlalchemy.orm import Session

from app.models.model_node import ModelNode
from app.models.model_version import ModelVersion
from app.models.orientation_rule import OrientationRule, OrientationCategory
from app.models.department import Department


class ModelVersionExportService:
    """模型版本导出服务"""
    
    def __init__(self, db: Session):
        self.db = db
        # 规则ID到Sheet名称的映射
        self._rule_sheet_map: Dict[int, str] = {}
    
    def export_to_excel(self, version_id: int) -> BytesIO:
        """
        导出模型版本结构到Excel
        
        Args:
            version_id: 模型版本ID
            
        Returns:
            BytesIO: Excel文件的字节流
        """
        # 获取版本信息
        version = self.db.query(ModelVersion).filter(ModelVersion.id == version_id).first()
        if not version:
            raise ValueError("模型版本不存在")
        
        # 获取所有根节点
        root_nodes = self.db.query(ModelNode).filter(
            ModelNode.version_id == version_id,
            ModelNode.parent_id.is_(None)
        ).order_by(ModelNode.sort_order).all()
        
        # 递归加载所有节点并收集导向规则ID
        orientation_rule_ids: Set[int] = set()
        for node in root_nodes:
            self._load_children(node, orientation_rule_ids)
        
        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "模型结构"
        
        # 先创建导向规则Sheet并建立映射
        self._rule_sheet_map.clear()
        if orientation_rule_ids:
            rules = self.db.query(OrientationRule).filter(
                OrientationRule.id.in_(orientation_rule_ids)
            ).order_by(OrientationRule.id).all()
            
            for rule in rules:
                # Sheet名称最多31个字符，截断并确保唯一
                sheet_name = self._get_safe_sheet_name(rule.name, wb)
                self._rule_sheet_map[rule.id] = sheet_name
                rule_ws = wb.create_sheet(title=sheet_name)
                self._write_orientation_rule_sheet(rule_ws, rule)
        
        # 写入模型结构Sheet（此时已有规则到Sheet的映射）
        self._write_model_structure_sheet(ws, version, root_nodes)
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _get_safe_sheet_name(self, name: str, wb: Workbook) -> str:
        """获取安全的Sheet名称（最多31字符，不重复）"""
        # Excel Sheet名称最多31个字符
        safe_name = name[:31] if len(name) > 31 else name
        # 替换非法字符
        for char in ['\\', '/', '*', '?', ':', '[', ']']:
            safe_name = safe_name.replace(char, '_')
        
        # 确保名称唯一
        existing_names = [ws.title for ws in wb.worksheets]
        if safe_name not in existing_names:
            return safe_name
        
        # 添加序号
        counter = 1
        while True:
            suffix = f"_{counter}"
            truncated_name = safe_name[:31-len(suffix)] + suffix
            if truncated_name not in existing_names:
                return truncated_name
            counter += 1
    
    def _write_model_structure_sheet(self, ws, version: ModelVersion, root_nodes: List[ModelNode]):
        """写入模型结构Sheet"""
        # 设置列宽
        ws.column_dimensions['A'].width = 50   # 节点名称
        ws.column_dimensions['B'].width = 30   # 节点编码
        ws.column_dimensions['C'].width = 12   # 算法类型
        ws.column_dimensions['D'].width = 15   # 权重/单价
        ws.column_dimensions['E'].width = 30   # 业务导向
        ws.column_dimensions['F'].width = 50   # 规则说明
        
        # 样式定义
        title_font = Font(name='微软雅黑', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题行
        title_text = f"评估模型结构 - {version.name} ({version.version})"
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        # 表头
        headers = ['节点名称', '节点编码', '算法类型', '权重/单价', '业务导向', '规则说明']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        ws.row_dimensions[2].height = 25
        
        # 写入数据
        row = 3
        for node in root_nodes:
            row = self._write_node(ws, node, row, 0, border)
    
    def _write_orientation_rule_sheet(self, ws, rule: OrientationRule):
        """写入导向规则Sheet"""
        # 样式定义
        title_font = Font(name='微软雅黑', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        section_font = Font(name='微软雅黑', size=12, bold=True, color='2c3e50')
        
        header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        label_font = Font(name='微软雅黑', size=10, bold=True)
        label_fill = PatternFill(start_color='e8f4f8', end_color='e8f4f8', fill_type='solid')
        
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        data_alignment_center = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置列宽（与基准表格一致，共7列）
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12
        
        row = 1
        
        # 标题（合并7列）
        title_cell = ws.cell(row=row, column=1, value=rule.name)
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        ws.merge_cells(f'A{row}:G{row}')
        ws.row_dimensions[row].height = 35
        row += 2
        
        # 基本信息部分
        section_cell = ws.cell(row=row, column=1, value="基本信息")
        section_cell.font = section_font
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        # 导向类别中文映射
        category_map = {
            OrientationCategory.benchmark_ladder: "基准阶梯",
            OrientationCategory.direct_ladder: "直接阶梯",
            OrientationCategory.other: "其他",
        }
        
        # 基本信息表格（第一列合并A-B，第二列合并C-G）
        info_items = [
            ("导向名称", rule.name),
            ("导向类别", category_map.get(rule.category, str(rule.category))),
        ]
        if rule.description:
            info_items.append(("导向规则描述", rule.description))
        info_items.extend([
            ("创建时间", rule.created_at.strftime('%Y-%m-%d %H:%M:%S')),
            ("更新时间", rule.updated_at.strftime('%Y-%m-%d %H:%M:%S')),
        ])
        
        for label, value in info_items:
            # 标签列（合并A-B）
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = label_font
            label_cell.fill = label_fill
            label_cell.border = border
            label_cell.alignment = data_alignment
            ws.merge_cells(f'A{row}:B{row}')
            # 为合并区域的第二个单元格也设置边框
            ws.cell(row=row, column=2).border = border
            
            # 值列（合并C-G）
            value_cell = ws.cell(row=row, column=3, value=value)
            value_cell.font = data_font
            value_cell.border = border
            value_cell.alignment = data_alignment
            ws.merge_cells(f'C{row}:G{row}')
            # 为合并区域的其他单元格也设置边框
            for col in range(4, 8):
                ws.cell(row=row, column=col).border = border
            
            # 根据内容计算行高（C-G合并后约65字符宽度）
            value_str = str(value) if value else ''
            row_height = self._calculate_row_height(value_str, column_width=65)
            ws.row_dimensions[row].height = row_height
            row += 1
        
        row += 1  # 空行
        
        # 根据类别显示关联数据
        if rule.category == OrientationCategory.benchmark_ladder and rule.benchmarks:
            # 导向基准部分
            section_cell = ws.cell(row=row, column=1, value="导向基准")
            section_cell.font = section_font
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
            
            # 基准类别中文映射
            benchmark_type_map = {
                "average": "平均值",
                "median": "中位数",
                "max": "最大值",
                "min": "最小值",
                "other": "其他",
            }
            
            # 表头
            benchmark_headers = ['科室代码', '科室名称', '基准类别', '管控力度', '统计开始时间', '统计结束时间', '基准值']
            for col_idx, header in enumerate(benchmark_headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            ws.row_dimensions[row].height = 25
            row += 1
            
            # 获取科室排序映射（按科室对照表的sort_order排序）
            dept_codes = [b.department_code for b in rule.benchmarks]
            dept_sort_map = {}
            if dept_codes:
                depts = self.db.query(Department).filter(
                    Department.his_code.in_(dept_codes)
                ).all()
                dept_sort_map = {d.his_code: float(d.sort_order) if d.sort_order else 999999 for d in depts}
            
            # 按科室对照表排序
            sorted_benchmarks = sorted(
                rule.benchmarks,
                key=lambda b: dept_sort_map.get(b.department_code, 999999)
            )
            
            # 数据行
            for benchmark in sorted_benchmarks:
                ws.cell(row=row, column=1, value=benchmark.department_code).border = border
                ws.cell(row=row, column=2, value=benchmark.department_name).border = border
                ws.cell(row=row, column=3, value=benchmark_type_map.get(benchmark.benchmark_type, benchmark.benchmark_type)).border = border
                ws.cell(row=row, column=4, value=float(benchmark.control_intensity)).border = border
                ws.cell(row=row, column=5, value=benchmark.stat_start_date.strftime('%Y-%m-%d')).border = border
                ws.cell(row=row, column=6, value=benchmark.stat_end_date.strftime('%Y-%m-%d')).border = border
                ws.cell(row=row, column=7, value=float(benchmark.benchmark_value)).border = border
                
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = data_alignment_center
                    # 数值格式
                    if col in [4, 7]:
                        cell.number_format = '0.0000'
                
                ws.row_dimensions[row].height = 20
                row += 1
            
            row += 1  # 空行
        
        # 导向阶梯
        if rule.category in [OrientationCategory.benchmark_ladder, OrientationCategory.direct_ladder] and rule.ladders:
            section_cell = ws.cell(row=row, column=1, value="导向阶梯")
            section_cell.font = section_font
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
            
            # 表头（4列，但要与7列对齐，每列合并部分单元格）
            # 阶梯次序(A-B), 阶梯下限(C-D), 阶梯上限(E-F), 调整力度(G)
            ladder_headers = [
                ('阶梯次序', 'A', 'B'),
                ('阶梯下限', 'C', 'D'),
                ('阶梯上限', 'E', 'F'),
                ('调整力度', 'G', 'G'),
            ]
            for header, start_col, end_col in ladder_headers:
                start_idx = ord(start_col) - ord('A') + 1
                cell = ws.cell(row=row, column=start_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                if start_col != end_col:
                    ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
                    end_idx = ord(end_col) - ord('A') + 1
                    ws.cell(row=row, column=end_idx).border = border
            ws.row_dimensions[row].height = 25
            row += 1
            
            # 数据行
            for ladder in sorted(rule.ladders, key=lambda x: x.ladder_order):
                lower = "-∞" if ladder.lower_limit is None else float(ladder.lower_limit)
                upper = "+∞" if ladder.upper_limit is None else float(ladder.upper_limit)
                
                # 阶梯次序(A-B)
                ws.cell(row=row, column=1, value=ladder.ladder_order).border = border
                ws.merge_cells(f'A{row}:B{row}')
                ws.cell(row=row, column=2).border = border
                
                # 阶梯下限(C-D)
                ws.cell(row=row, column=3, value=lower).border = border
                ws.merge_cells(f'C{row}:D{row}')
                ws.cell(row=row, column=4).border = border
                
                # 阶梯上限(E-F)
                ws.cell(row=row, column=5, value=upper).border = border
                ws.merge_cells(f'E{row}:F{row}')
                ws.cell(row=row, column=6).border = border
                
                # 调整力度(G)
                ws.cell(row=row, column=7, value=float(ladder.adjustment_intensity)).border = border
                
                for col in [1, 3, 5, 7]:
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = data_alignment_center
                    # 数值格式
                    if col in [3, 5, 7] and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0000'
                
                ws.row_dimensions[row].height = 20
                row += 1
    
    def _load_children(self, node: ModelNode, orientation_rule_ids: Set[int]):
        """递归加载子节点并收集导向规则ID"""
        # 收集当前节点的导向规则ID
        if node.orientation_rule_ids:
            orientation_rule_ids.update(node.orientation_rule_ids)
        
        children = self.db.query(ModelNode).filter(
            ModelNode.parent_id == node.id
        ).order_by(ModelNode.sort_order).all()
        node.children = children
        for child in children:
            self._load_children(child, orientation_rule_ids)
    
    def _get_orientation_rule_names(self, rule_ids: Optional[List[int]]) -> str:
        """获取导向规则名称列表（纯文本，用于显示）"""
        if not rule_ids:
            return '-'
        rules = self.db.query(OrientationRule).filter(
            OrientationRule.id.in_(rule_ids)
        ).all()
        if not rules:
            return '-'
        return ', '.join([rule.name for rule in rules])
    
    def _get_orientation_rules_with_links(self, rule_ids: Optional[List[int]]) -> List[Tuple[int, str, str]]:
        """
        获取导向规则信息列表（包含ID、名称和Sheet名称）
        
        Returns:
            List of (rule_id, rule_name, sheet_name) tuples
        """
        if not rule_ids:
            return []
        rules = self.db.query(OrientationRule).filter(
            OrientationRule.id.in_(rule_ids)
        ).all()
        result = []
        for rule in rules:
            sheet_name = self._rule_sheet_map.get(rule.id, '')
            if sheet_name:
                result.append((rule.id, rule.name, sheet_name))
        return result
    
    def _format_weight(self, weight: Optional[float], unit: Optional[str]) -> str:
        """格式化权重显示"""
        if weight is None:
            return '-'
        unit_value = unit or '%'
        # 如果单位是百分比，将值乘以100显示
        display_value = weight * 100 if unit_value == '%' else weight
        return f"{display_value:.2f} {unit_value}"
    
    def _get_calc_type_display(self, node: ModelNode) -> str:
        """获取算法类型显示文本"""
        if not node.is_leaf:
            return '-'
        if node.calc_type == 'statistical':
            return '指标'
        elif node.calc_type == 'calculational':
            return '目录'
        return '-'
    
    def _get_node_type_tag(self, node: ModelNode) -> str:
        """获取节点类型标签"""
        tags = []
        if node.node_type == 'sequence':
            tags.append('[序列]')
        else:
            tags.append('[维度]')
        if node.is_leaf:
            tags.append('[末级]')
        return ' '.join(tags)
    
    def _write_node(self, ws, node: ModelNode, row: int, level: int, border: Border) -> int:
        """递归写入节点数据"""
        # 数据样式
        data_font = Font(name='微软雅黑', size=10)
        data_font_bold = Font(name='微软雅黑', size=10, bold=True)
        link_font = Font(name='微软雅黑', size=10, color='0563C1', underline='single')
        data_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        data_alignment_center = Alignment(horizontal='center', vertical='center')
        data_alignment_right = Alignment(horizontal='right', vertical='center')
        
        # 根据层级设置不同的背景色
        level_fills = [
            PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),  # 根节点
            PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),  # 一级
            None,  # 二级及以下无背景
        ]
        fill = level_fills[min(level, 2)]
        
        # 计算缩进
        indent = '    ' * level
        
        # 节点名称（带类型标签）
        type_tag = self._get_node_type_tag(node)
        name_text = f"{indent}{node.name} {type_tag}"
        
        # 写入数据
        ws.cell(row=row, column=1, value=name_text)
        ws.cell(row=row, column=2, value=node.code)
        ws.cell(row=row, column=3, value=self._get_calc_type_display(node))
        ws.cell(row=row, column=4, value=self._format_weight(node.weight, node.unit) if node.is_leaf else '-')
        
        # 业务导向列：添加超链接
        orientation_cell = ws.cell(row=row, column=5)
        if node.is_leaf and node.orientation_rule_ids:
            rules_info = self._get_orientation_rules_with_links(node.orientation_rule_ids)
            if rules_info:
                # 只有一个规则时，直接添加超链接
                if len(rules_info) == 1:
                    rule_id, rule_name, sheet_name = rules_info[0]
                    orientation_cell.value = rule_name
                    orientation_cell.hyperlink = f"#'{sheet_name}'!A1"
                    orientation_cell.font = link_font
                else:
                    # 多个规则时，显示名称列表（第一个带超链接）
                    # Excel单元格只能有一个超链接，所以显示所有名称，链接到第一个
                    names = [info[1] for info in rules_info]
                    orientation_cell.value = ', '.join(names)
                    # 链接到第一个规则的Sheet
                    orientation_cell.hyperlink = f"#'{rules_info[0][2]}'!A1"
                    orientation_cell.font = link_font
            else:
                orientation_cell.value = '-'
        else:
            orientation_cell.value = '-'
        
        rule_text = node.rule or '-'
        ws.cell(row=row, column=6, value=rule_text)
        
        # 应用样式
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if col != 5:  # 业务导向列已单独设置字体
                cell.font = data_font_bold if level == 0 else data_font
            elif not (node.is_leaf and node.orientation_rule_ids and self._get_orientation_rules_with_links(node.orientation_rule_ids)):
                # 业务导向列无链接时使用普通字体
                cell.font = data_font_bold if level == 0 else data_font
            cell.border = border
            if fill:
                cell.fill = fill
            
            if col == 1:  # 节点名称
                cell.alignment = data_alignment_left
            elif col in [3, 4, 5]:  # 算法类型、权重、业务导向
                cell.alignment = data_alignment_center
            elif col == 6:  # 规则说明
                cell.alignment = data_alignment_left
            else:
                cell.alignment = data_alignment_left
        
        # 根据规则说明内容计算行高
        row_height = self._calculate_row_height(rule_text, column_width=50)
        ws.row_dimensions[row].height = row_height
        row += 1
        
        # 递归处理子节点
        if hasattr(node, 'children') and node.children:
            for child in node.children:
                row = self._write_node(ws, child, row, level + 1, border)
        
        return row
    
    def _calculate_row_height(self, text: str, column_width: int = 50) -> float:
        """
        根据文本内容计算行高
        
        Args:
            text: 单元格文本内容
            column_width: 列宽（字符数）
            
        Returns:
            行高（磅）
        """
        if not text or text == '-':
            return 22  # 默认行高
        
        # 计算文本长度（中文字符算2个字符宽度）
        text_length = 0
        for char in text:
            if '\u4e00' <= char <= '\u9fff':
                text_length += 2  # 中文字符
            else:
                text_length += 1  # 英文字符
        
        # 计算需要的行数（考虑换行）
        lines_from_length = max(1, (text_length + column_width - 1) // column_width)
        
        # 计算文本中的换行符数量
        newline_count = text.count('\n')
        
        # 取较大值
        total_lines = max(lines_from_length, newline_count + 1)
        
        # 每行约15磅，最小22磅，最大200磅
        row_height = max(22, min(200, total_lines * 15 + 7))
        
        return row_height
