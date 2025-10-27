"""
维度目录智能导入服务
"""
from typing import Dict, List, Any, Optional, Tuple
from io import BytesIO
import openpyxl
from sqlalchemy.orm import Session
from sqlalchemy import or_
from difflib import SequenceMatcher
import uuid

from app.models.charge_item import ChargeItem
from app.models.model_node import ModelNode
from app.models.dimension_item_mapping import DimensionItemMapping


class DimensionImportService:
    """维度目录智能导入服务"""
    
    # 会话存储（生产环境应使用Redis）
    _sessions: Dict[str, Dict[str, Any]] = {}
    
    @classmethod
    def parse_excel(cls, file_content: bytes, sheet_name: Optional[str] = None, skip_rows: int = 0) -> Dict[str, Any]:
        """
        第一步：解析Excel文件，返回列名和预览数据
        
        Args:
            file_content: Excel文件内容
            sheet_name: 工作表名称（可选，默认使用第一个sheet）
            skip_rows: 跳过前N行（默认0）
        
        Returns:
            {
                "session_id": "uuid",
                "sheet_names": ["Sheet1", "Sheet2", ...],
                "current_sheet": "Sheet1",
                "headers": ["列1", "列2", ...],
                "preview_data": [[值1, 值2, ...], ...],
                "total_rows": 100,
                "skip_rows": 0,
                "suggested_mapping": {
                    "item_code": "收费编码",
                    "dimension_plan": "维度预案",
                    "expert_opinion": "专家意见"
                }
            }
        """
        # 验证文件内容
        if not file_content or len(file_content) == 0:
            raise ValueError("文件内容为空")
        
        # 读取Excel
        try:
            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        except Exception as e:
            error_msg = str(e)
            if "zip" in error_msg.lower():
                raise ValueError("文件格式错误，请上传有效的 Excel 文件（.xlsx 格式）")
            else:
                raise ValueError(f"无法读取 Excel 文件: {error_msg}")
        
        # 获取所有sheet名称
        sheet_names = wb.sheetnames
        
        # 选择工作表
        if sheet_name and sheet_name in sheet_names:
            ws = wb[sheet_name]
            current_sheet = sheet_name
        else:
            ws = wb.active
            current_sheet = ws.title
        
        # 获取所有行
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel文件为空")
        
        # 跳过前N行
        if skip_rows > 0:
            if skip_rows >= len(rows):
                raise ValueError(f"跳过行数({skip_rows})超过总行数({len(rows)})")
            rows = rows[skip_rows:]
        
        if not rows:
            raise ValueError("跳过指定行数后没有剩余数据")
        
        # 第一行作为表头
        def get_excel_column_name(col_index):
            """将列索引转换为Excel列名（A, B, C, ..., Z, AA, AB, ...）"""
            result = ""
            while col_index >= 0:
                result = chr(col_index % 26 + 65) + result
                col_index = col_index // 26 - 1
            return result
        
        headers = []
        for i, cell in enumerate(rows[0]):
            col_name = get_excel_column_name(i)
            if cell is not None and str(cell).strip():
                # 格式：列名 (Excel列)，例如：收费编码 (A)
                headers.append(f"{str(cell).strip()} ({col_name})")
            else:
                # 格式：(空列-Excel列)，例如：(空列-D)
                headers.append(f"(空列-{col_name})")
        
        # 数据行（跳过表头）
        data_rows = rows[1:]
        total_rows = len(data_rows)
        
        if total_rows == 0:
            raise ValueError("Excel文件没有数据行")
        
        # 预览数据（前10行）
        preview_data = []
        for row in data_rows[:10]:
            preview_data.append([str(cell).strip() if cell is not None else "" 
                               for cell in row])
        
        # 建议的字段映射
        suggested_mapping = cls._suggest_field_mapping(headers)
        
        # 生成会话ID并存储数据
        session_id = str(uuid.uuid4())
        cls._sessions[session_id] = {
            "file_content": file_content,
            "sheet_name": current_sheet,
            "skip_rows": skip_rows,
            "headers": headers,
            "total_rows": total_rows
        }
        
        return {
            "session_id": session_id,
            "sheet_names": sheet_names,
            "current_sheet": current_sheet,
            "headers": headers,
            "preview_data": preview_data,
            "total_rows": total_rows,
            "skip_rows": skip_rows,
            "suggested_mapping": suggested_mapping
        }
    
    @classmethod
    def _suggest_field_mapping(cls, headers: List[str]) -> Dict[str, str]:
        """
        根据表头建议字段映射
        
        Args:
            headers: Excel表头列表
        
        Returns:
            建议的映射关系
        """
        suggested = {}
        
        # 定义关键词映射规则
        keywords_map = {
            "item_code": ["收费编码", "项目编码", "编码", "代码", "sfbm", "xmbm"],
            "dimension_plan": ["维度预案", "维度", "预案", "分类"],
            "expert_opinion": ["专家意见", "专家", "意见", "建议"]
        }
        
        for header in headers:
            header_lower = header.lower()
            
            # 尝试匹配每个字段
            for field, keywords in keywords_map.items():
                for keyword in keywords:
                    if keyword.lower() in header_lower:
                        suggested[field] = header
                        break
                if field in suggested:
                    break
        
        return suggested
    
    @classmethod
    def extract_unique_values(
        cls,
        session_id: str,
        field_mapping: Dict[str, str],
        model_version_id: int,
        db: Session,
        match_by: str = "code"
    ) -> Dict[str, Any]:
        """
        第二步：提取维度预案和专家意见的唯一值，并提供智能匹配建议
        
        Args:
            session_id: 会话ID
            field_mapping: 字段映射 {"item_code": "收费编码", ...}
            model_version_id: 模型版本ID
            db: 数据库会话
            match_by: 匹配方式，"code"(按编码) 或 "name"(按名称)
        
        Returns:
            {
                "unique_values": [
                    {
                        "value": "4D",
                        "source": "expert_opinion",
                        "count": 120,
                        "suggested_dimensions": [...]
                    },
                    ...
                ],
                "system_dimensions": [...]
            }
        """
        # 获取会话数据
        session_data = cls._sessions.get(session_id)
        if not session_data:
            raise ValueError("会话已过期或不存在")
        
        # 验证字段映射
        if "item_code" not in field_mapping:
            raise ValueError("必须映射收费编码字段")
        
        # 读取Excel数据
        file_content = session_data["file_content"]
        headers = session_data["headers"]
        sheet_name = session_data.get("sheet_name")
        skip_rows = session_data.get("skip_rows", 0)
        
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        
        # 跳过前N行
        if skip_rows > 0:
            rows = rows[skip_rows:]
        
        data_rows = rows[1:]  # 跳过表头
        
        # 提取唯一值
        unique_values_dict = {}
        
        for source_field in ["dimension_plan", "expert_opinion"]:
            if source_field not in field_mapping:
                continue
            
            col_name = field_mapping[source_field]
            if col_name not in headers:
                continue
            
            col_idx = headers.index(col_name)
            
            for row in data_rows:
                if col_idx < len(row) and row[col_idx]:
                    value = str(row[col_idx]).strip()
                    if value:
                        key = (value, source_field)
                        if key not in unique_values_dict:
                            unique_values_dict[key] = 0
                        unique_values_dict[key] += 1
        
        # 获取系统维度列表（只获取叶子节点）
        system_dimensions = cls._get_system_dimensions(model_version_id, db)
        
        # 为每个唯一值提供智能匹配建议
        unique_values = []
        for (value, source), count in unique_values_dict.items():
            suggested_dims = cls._suggest_dimensions(value, system_dimensions)
            unique_values.append({
                "value": value,
                "source": source,
                "count": count,
                "suggested_dimensions": suggested_dims
            })
        
        # 按来源分组，然后按出现次数排序
        # 先按来源排序（expert_opinion优先），再按出现次数排序
        unique_values.sort(key=lambda x: (0 if x["source"] == "expert_opinion" else 1, -x["count"]))
        
        # 保存到会话
        cls._sessions[session_id]["field_mapping"] = field_mapping
        cls._sessions[session_id]["unique_values"] = unique_values
        cls._sessions[session_id]["match_by"] = match_by  # 保存匹配方式
        
        return {
            "unique_values": unique_values,
            "system_dimensions": system_dimensions
        }
    
    @classmethod
    def _get_system_dimensions(cls, model_version_id: int, db: Session) -> List[Dict[str, Any]]:
        """获取系统维度列表（只获取叶子节点）"""
        nodes = db.query(ModelNode).filter(
            ModelNode.version_id == model_version_id,
            ModelNode.is_leaf == True
        ).all()
        
        dimensions = []
        for node in nodes:
            # 构建完整路径
            full_path = cls._build_node_path(node, db)
            dimensions.append({
                "id": node.id,
                "name": node.name,
                "code": node.code,
                "full_path": full_path
            })
        
        return dimensions
    
    @classmethod
    def _build_node_path(cls, node: ModelNode, db: Session) -> str:
        """构建节点的完整路径"""
        path_parts = [node.name]
        current = node
        
        while current.parent_id:
            parent = db.query(ModelNode).filter(ModelNode.id == current.parent_id).first()
            if parent:
                path_parts.insert(0, parent.name)
                current = parent
            else:
                break
        
        return " > ".join(path_parts)
    
    @classmethod
    def _suggest_dimensions(
        cls,
        value: str,
        system_dimensions: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        为给定值建议匹配的维度
        
        使用模糊匹配算法
        """
        suggestions = []
        value_lower = value.lower()
        
        for dim in system_dimensions:
            # 计算相似度
            name_similarity = SequenceMatcher(None, value_lower, dim["name"].lower()).ratio()
            code_similarity = SequenceMatcher(None, value_lower, dim["code"].lower()).ratio()
            
            # 检查是否包含
            contains_in_name = value_lower in dim["name"].lower()
            contains_in_path = value_lower in dim["full_path"].lower()
            
            # 计算综合得分
            score = max(name_similarity, code_similarity)
            if contains_in_name:
                score += 0.3
            if contains_in_path:
                score += 0.2
            
            if score > 0.3:  # 阈值
                suggestions.append({
                    "id": dim["id"],
                    "name": dim["name"],
                    "code": dim["code"],
                    "full_path": dim["full_path"],
                    "score": score
                })
        
        # 按得分排序
        suggestions.sort(key=lambda x: x["score"], reverse=True)
        
        # 返回前5个建议
        return suggestions[:5]
    
    @classmethod
    def generate_preview(
        cls,
        session_id: str,
        value_mapping: List[Dict[str, Any]],
        db: Session
    ) -> Dict[str, Any]:
        """
        第三步：生成导入预览
        
        Args:
            session_id: 会话ID
            value_mapping: 维度值映射 [{"value": "4D", "source": "expert_opinion", "dimension_codes": ["D001", "D002"]}, ...]
            db: 数据库会话
        
        Returns:
            {
                "preview_items": [...],
                "statistics": {
                    "total": 500,
                    "ok": 450,
                    "warning": 48,
                    "error": 2
                }
            }
        """
        # 获取会话数据
        session_data = cls._sessions.get(session_id)
        if not session_data:
            raise ValueError("会话已过期或不存在")
        
        field_mapping = session_data.get("field_mapping")
        if not field_mapping:
            raise ValueError("缺少字段映射信息")
        
        # 构建值到维度的映射字典
        value_to_dims = {}
        for mapping in value_mapping:
            # 如果是Pydantic模型，使用属性访问；如果是字典，使用字典访问
            if hasattr(mapping, 'value'):
                key = (mapping.value, mapping.source)
                value_to_dims[key] = mapping.dimension_codes if mapping.dimension_codes else []
            else:
                key = (mapping["value"], mapping["source"])
                value_to_dims[key] = mapping.get("dimension_codes", [])
        
        # 读取Excel数据
        file_content = session_data["file_content"]
        headers = session_data["headers"]
        sheet_name = session_data.get("sheet_name")
        skip_rows = session_data.get("skip_rows", 0)
        
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        
        # 跳过前N行
        if skip_rows > 0:
            rows = rows[skip_rows:]
        
        data_rows = rows[1:]  # 跳过表头
        
        # 获取列索引
        item_code_col = headers.index(field_mapping["item_code"])
        dimension_plan_col = headers.index(field_mapping["dimension_plan"]) if "dimension_plan" in field_mapping else -1
        expert_opinion_col = headers.index(field_mapping["expert_opinion"]) if "expert_opinion" in field_mapping else -1
        
        # 获取匹配方式
        match_by = session_data.get("match_by", "code")
        
        # 获取所有收费项目（用于验证和转换）
        all_charge_items_by_code = {item.item_code: item for item in db.query(ChargeItem).all()}
        all_charge_items_by_name = {item.item_name: item for item in db.query(ChargeItem).all()}
        
        # 获取所有维度（用于验证）
        all_dimensions = {node.code: node for node in db.query(ModelNode).all()}
        
        # 获取已存在的映射关系
        existing_mappings = set()
        for mapping in db.query(DimensionItemMapping).all():
            existing_mappings.add((mapping.dimension_code, mapping.item_code))
        
        # 生成预览数据
        preview_items = []
        statistics = {"total": 0, "ok": 0, "warning": 0, "error": 0}
        
        for row in data_rows:
            if item_code_col >= len(row):
                continue
            
            item_value = str(row[item_code_col]).strip() if row[item_code_col] else ""
            if not item_value:
                continue
            
            # 根据匹配方式获取收费项目编码
            if match_by == "name":
                # 按名称匹配，需要转换为编码
                charge_item = all_charge_items_by_name.get(item_value)
                if charge_item:
                    item_code = charge_item.item_code
                    item_name = charge_item.item_name
                else:
                    # 名称不存在，使用原值作为编码（会在后续标记为warning）
                    item_code = item_value
                    item_name = ""
            else:
                # 按编码匹配
                item_code = item_value
                charge_item = all_charge_items_by_code.get(item_code)
                item_name = charge_item.item_name if charge_item else ""
            
            # 获取维度预案和专家意见
            dimension_plan = ""
            expert_opinion = ""
            
            if dimension_plan_col >= 0 and dimension_plan_col < len(row):
                dimension_plan = str(row[dimension_plan_col]).strip() if row[dimension_plan_col] else ""
            
            if expert_opinion_col >= 0 and expert_opinion_col < len(row):
                expert_opinion = str(row[expert_opinion_col]).strip() if row[expert_opinion_col] else ""
            
            # 确定使用哪个值（专家意见优先）
            source_value = expert_opinion if expert_opinion else dimension_plan
            source_type = "expert_opinion" if expert_opinion else "dimension_plan"
            
            if not source_value:
                continue
            
            # 获取对应的维度编码列表
            dimension_codes = value_to_dims.get((source_value, source_type), [])
            
            # 如果没有指定维度，跳过这条记录（用户可以选择不导入某些值）
            if not dimension_codes:
                continue
            
            # 为每个维度编码生成一条预览记录
            for dim_code in dimension_codes:
                statistics["total"] += 1
                
                # 验证收费项目是否存在（使用编码验证）
                charge_item = all_charge_items_by_code.get(item_code)
                if not item_name and charge_item:
                    item_name = charge_item.item_name
                
                # 验证维度是否存在
                dimension = all_dimensions.get(dim_code)
                if not dimension:
                    preview_items.append({
                        "item_code": item_code,
                        "item_name": item_name,
                        "dimension_code": dim_code,
                        "dimension_name": "",
                        "dimension_path": "",
                        "source": source_type,
                        "source_value": source_value,
                        "status": "error",
                        "message": "目标维度不存在"
                    })
                    statistics["error"] += 1
                    continue
                
                dimension_path = cls._build_node_path(dimension, db)
                
                # 检查状态
                status = "ok"
                message = ""
                
                if not charge_item:
                    status = "warning"
                    if match_by == "name":
                        message = f"收费项目名称'{item_value}'在系统中不存在"
                    else:
                        message = "收费项目编码在系统中不存在"
                    statistics["warning"] += 1
                elif (dim_code, item_code) in existing_mappings:
                    status = "ok"  # 改为ok状态，因为会覆盖
                    message = "该收费项目已存在，将被覆盖"
                    statistics["ok"] += 1
                else:
                    statistics["ok"] += 1
                
                preview_items.append({
                    "item_code": item_code,
                    "item_name": item_name,
                    "dimension_code": dim_code,
                    "dimension_name": dimension.name,
                    "dimension_path": dimension_path,
                    "source": source_type,
                    "source_value": source_value,
                    "status": status,
                    "message": message
                })
        
        # 保存到会话
        cls._sessions[session_id]["preview_items"] = preview_items
        
        return {
            "preview_items": preview_items,
            "statistics": statistics
        }
    
    @classmethod
    def execute_import(
        cls,
        session_id: str,
        confirmed_items: Optional[List[Dict[str, Any]]],
        db: Session
    ) -> Dict[str, Any]:
        """
        执行导入
        
        Args:
            session_id: 会话ID
            confirmed_items: 用户确认的导入项（如果为空则导入所有预览项）
            db: 数据库会话
        
        Returns:
            {
                "success": True,
                "report": {
                    "success_count": 450,
                    "skipped_count": 48,
                    "error_count": 2,
                    "errors": [...]
                }
            }
        """
        # 获取会话数据
        session_data = cls._sessions.get(session_id)
        if not session_data:
            raise ValueError("会话已过期或不存在")
        
        preview_items = session_data.get("preview_items")
        if not preview_items:
            raise ValueError("缺少预览数据")
        
        # 如果用户指定了确认项，则只导入这些项
        items_to_import = confirmed_items if confirmed_items else preview_items
        
        success_count = 0
        skipped_count = 0
        error_count = 0
        errors = []
        
        for item in items_to_import:
            try:
                # 跳过错误状态的项
                if item.get("status") == "error":
                    error_count += 1
                    errors.append({
                        "item_code": item["item_code"],
                        "dimension_code": item["dimension_code"],
                        "reason": item.get("message", "未知错误")
                    })
                    continue
                
                # 检查是否已存在映射
                existing = db.query(DimensionItemMapping).filter(
                    DimensionItemMapping.dimension_code == item["dimension_code"],
                    DimensionItemMapping.item_code == item["item_code"]
                ).first()
                
                if existing:
                    # 删除旧记录，实现覆盖效果
                    db.delete(existing)
                    db.flush()
                
                # 创建新映射（即使收费项目不存在也允许创建）
                mapping = DimensionItemMapping(
                    dimension_code=item["dimension_code"],
                    item_code=item["item_code"]
                )
                db.add(mapping)
                db.flush()  # 立即执行以检测错误
                success_count += 1
                
            except Exception as e:
                db.rollback()  # 回滚当前记录
                error_count += 1
                errors.append({
                    "item_code": item.get("item_code", "unknown"),
                    "dimension_code": item.get("dimension_code", ""),
                    "reason": str(e)
                })
        
        # 提交所有成功的记录
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            raise ValueError(f"提交事务失败: {str(e)}")
        
        # 清理会话
        if session_id in cls._sessions:
            del cls._sessions[session_id]
        
        return {
            "success": True,
            "report": {
                "success_count": success_count,
                "skipped_count": skipped_count,
                "error_count": error_count,
                "errors": errors[:100]  # 最多返回100条错误
            }
        }
