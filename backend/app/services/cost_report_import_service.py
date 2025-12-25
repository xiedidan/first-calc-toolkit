"""
成本报表智能导入服务
"""
from typing import Dict, List, Any, Optional
from io import BytesIO
from decimal import Decimal
import openpyxl
from sqlalchemy.orm import Session
from difflib import SequenceMatcher
import uuid
import re

from app.models.cost_report import CostReport
from app.models.department import Department


class CostReportImportService:
    """成本报表智能导入服务"""
    
    # 会话存储（生产环境应使用Redis）
    _sessions: Dict[str, Dict[str, Any]] = {}
    
    @classmethod
    def parse_excel(cls, file_content: bytes, sheet_name: Optional[str] = None, skip_rows: int = 0, header_row: int = 1) -> Dict[str, Any]:
        """
        第一步：解析Excel文件，返回列名和预览数据
        
        Args:
            file_content: Excel文件内容
            sheet_name: 工作表名称
            skip_rows: 跳过前N行
            header_row: 标题行位置（从1开始，相对于跳过行数后的位置）
        """
        if not file_content or len(file_content) == 0:
            raise ValueError("文件内容为空")
        
        try:
            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        except Exception as e:
            error_msg = str(e)
            if "zip" in error_msg.lower():
                raise ValueError("文件格式错误，请上传有效的 Excel 文件（.xlsx 格式）")
            else:
                raise ValueError(f"无法读取 Excel 文件: {error_msg}")
        
        sheet_names = wb.sheetnames
        
        if sheet_name and sheet_name in sheet_names:
            ws = wb[sheet_name]
            current_sheet = sheet_name
        else:
            ws = wb.active
            current_sheet = ws.title
        
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise ValueError("Excel文件为空")
        
        # 计算标题行的实际索引（0-based）
        # header_row 是从1开始的行号，表示Excel中的实际行号
        header_row_index = header_row - 1
        
        if header_row_index >= len(all_rows):
            raise ValueError(f"标题行位置({header_row})超过总行数({len(all_rows)})")
        
        def get_excel_column_name(col_index):
            result = ""
            while col_index >= 0:
                result = chr(col_index % 26 + 65) + result
                col_index = col_index // 26 - 1
            return result
        
        # 从标题行读取表头
        headers = []
        for i, cell in enumerate(all_rows[header_row_index]):
            col_name = get_excel_column_name(i)
            if cell is not None and str(cell).strip():
                headers.append(f"{str(cell).strip()} ({col_name})")
            else:
                headers.append(f"(空列-{col_name})")
        
        # 数据行从标题行之后开始
        data_rows = all_rows[header_row_index + 1:]
        total_rows = len(data_rows)
        
        if total_rows == 0:
            raise ValueError("Excel文件没有数据行")
        
        preview_data = []
        for row in data_rows[:10]:
            preview_data.append([str(cell).strip() if cell is not None else "" for cell in row])
        
        suggested_mapping = cls._suggest_field_mapping(headers)
        
        session_id = str(uuid.uuid4())
        cls._sessions[session_id] = {
            "file_content": file_content,
            "sheet_name": current_sheet,
            "skip_rows": skip_rows,
            "header_row": header_row,
            "headers": headers,
            "total_rows": total_rows
        }
        
        return {
            "session_id": session_id,
            "sheet_names": sheet_names,
            "current_sheet": current_sheet,
            "headers": headers,
            "preview_data": preview_data,
            "total_rows": total_rows,
            "skip_rows": skip_rows,
            "header_row": header_row,
            "suggested_mapping": suggested_mapping
        }
    
    @classmethod
    def _suggest_field_mapping(cls, headers: List[str]) -> Dict[str, str]:
        """根据表头建议字段映射"""
        suggested = {}
        
        keywords_map = {
            "period": ["年月", "月份", "期间", "日期"],
            "department_code": ["核算单元代码", "核算单元编码", "科室代码", "科室编码", "部门代码", "部门编码", "代码"],
            "department_name": ["核算单元名称", "核算单元", "科室名称", "科室", "部门名称", "部门"],
            "personnel_cost": ["人员经费", "人员费用", "人员成本", "工资", "薪酬"],
            "material_cost": ["卫生材料", "材料费", "不收费卫生材料", "材料成本"],
            "medicine_cost": ["药品费", "不收费药品", "药品成本", "药品"],
            "depreciation_cost": ["折旧", "折旧费", "折旧风险", "风险费"],
            "other_cost": ["其他费用", "其他成本", "其他"]
        }
        
        for header in headers:
            header_lower = header.lower()
            for field, keywords in keywords_map.items():
                if field in suggested:
                    continue
                for keyword in keywords:
                    if keyword.lower() in header_lower:
                        suggested[field] = header
                        break
        
        return suggested
    
    @classmethod
    def extract_unique_values(
        cls,
        session_id: str,
        field_mapping: Dict[str, str],
        db: Session,
        hospital_id: int,
        match_by: str = "code"
    ) -> Dict[str, Any]:
        """
        第二步：提取科室名称的唯一值，并提供智能匹配建议
        """
        session_data = cls._sessions.get(session_id)
        if not session_data:
            raise ValueError("会话已过期或不存在")
        
        file_content = session_data["file_content"]
        headers = session_data["headers"]
        sheet_name = session_data.get("sheet_name")
        header_row = session_data.get("header_row", 1)
        
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        all_rows = list(ws.iter_rows(values_only=True))
        # 数据行从标题行之后开始
        header_row_index = header_row - 1
        data_rows = all_rows[header_row_index + 1:]
        
        # 获取科室名称列索引
        dept_col_name = field_mapping.get("department_name", "")
        if not dept_col_name or dept_col_name not in headers:
            raise ValueError("必须映射科室名称字段")
        
        dept_col_idx = headers.index(dept_col_name)
        
        # 提取唯一值
        unique_values_dict = {}
        for row in data_rows:
            if dept_col_idx < len(row) and row[dept_col_idx]:
                # 清理空格和Tab
                value = str(row[dept_col_idx]).strip()
                value = re.sub(r'[\t\r\n]+', '', value)
                value = re.sub(r'\s+', ' ', value).strip()
                if value:
                    if value not in unique_values_dict:
                        unique_values_dict[value] = 0
                    unique_values_dict[value] += 1
        
        # 获取系统科室列表
        system_departments = cls._get_system_departments(db, hospital_id)
        
        # 为每个唯一值提供智能匹配建议
        unique_values = []
        for value, count in unique_values_dict.items():
            suggested_depts = cls._suggest_departments(value, system_departments)
            unique_values.append({
                "value": value,
                "count": count,
                "suggested_departments": suggested_depts
            })
        
        unique_values.sort(key=lambda x: -x["count"])
        
        cls._sessions[session_id]["field_mapping"] = field_mapping
        cls._sessions[session_id]["match_by"] = match_by
        
        return {
            "unique_values": unique_values,
            "system_departments": system_departments
        }
    
    @classmethod
    def _get_system_departments(cls, db: Session, hospital_id: int) -> List[Dict[str, Any]]:
        """获取系统科室列表（按核算单元去重）"""
        departments = db.query(Department).filter(
            Department.hospital_id == hospital_id
        ).all()
        
        # 按核算单元代码去重
        seen_codes = set()
        result = []
        for d in departments:
            code = d.accounting_unit_code or d.his_code
            name = d.accounting_unit_name or d.his_name
            if code and code not in seen_codes:
                seen_codes.add(code)
                result.append({
                    "id": d.id, 
                    "code": code, 
                    "name": name, 
                    "score": 1.0
                })
        
        return result
    
    @classmethod
    def _suggest_departments(cls, value: str, system_departments: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """为给定值建议匹配的科室"""
        suggestions = []
        value_lower = value.lower().strip()
        
        for dept in system_departments:
            name_similarity = SequenceMatcher(None, value_lower, dept["name"].lower()).ratio()
            code_similarity = SequenceMatcher(None, value_lower, dept["code"].lower()).ratio()
            
            contains_in_name = value_lower in dept["name"].lower() or dept["name"].lower() in value_lower
            
            score = max(name_similarity, code_similarity)
            if contains_in_name:
                score += 0.3
            
            if score > 0.3:
                suggestions.append({
                    "id": dept["id"],
                    "code": dept["code"],
                    "name": dept["name"],
                    "score": min(score, 1.0)
                })
        
        suggestions.sort(key=lambda x: x["score"], reverse=True)
        return suggestions[:5]
    
    @classmethod
    def generate_preview(
        cls,
        session_id: str,
        value_mapping: List[Dict[str, Any]],
        db: Session,
        hospital_id: int
    ) -> Dict[str, Any]:
        """
        第三步：生成导入预览
        """
        session_data = cls._sessions.get(session_id)
        if not session_data:
            raise ValueError("会话已过期或不存在")
        
        field_mapping = session_data.get("field_mapping")
        if not field_mapping:
            raise ValueError("缺少字段映射信息")
        
        match_by = session_data.get("match_by", "code")
        
        # 构建科室名称到代码的映射
        name_to_code = {}
        for mapping in value_mapping:
            if hasattr(mapping, 'value'):
                name_to_code[mapping.value] = mapping.department_code
            else:
                name_to_code[mapping["value"]] = mapping.get("department_code")
        
        # 读取Excel数据
        file_content = session_data["file_content"]
        headers = session_data["headers"]
        sheet_name = session_data.get("sheet_name")
        header_row = session_data.get("header_row", 1)
        
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        all_rows = list(ws.iter_rows(values_only=True))
        # 数据行从标题行之后开始
        header_row_index = header_row - 1
        data_rows = all_rows[header_row_index + 1:]
        
        # 获取列索引
        def get_col_idx(field_name):
            col_name = field_mapping.get(field_name, "")
            if col_name and col_name in headers:
                return headers.index(col_name)
            return -1
        
        period_col = get_col_idx("period")
        dept_code_col = get_col_idx("department_code")
        dept_name_col = get_col_idx("department_name")
        personnel_col = get_col_idx("personnel_cost")
        material_col = get_col_idx("material_cost")
        medicine_col = get_col_idx("medicine_cost")
        depreciation_col = get_col_idx("depreciation_cost")
        other_col = get_col_idx("other_cost")
        
        # 获取系统科室（按核算单元代码和名称索引）
        all_depts = db.query(Department).filter(Department.hospital_id == hospital_id).all()
        all_departments_by_code = {}
        all_departments_by_name = {}
        for d in all_depts:
            code = d.accounting_unit_code or d.his_code
            name = d.accounting_unit_name or d.his_name
            if code and code not in all_departments_by_code:
                all_departments_by_code[code] = d
            if name and name not in all_departments_by_name:
                all_departments_by_name[name] = d
        
        # 获取已存在的记录
        existing_records = set()
        for record in db.query(CostReport).filter(CostReport.hospital_id == hospital_id).all():
            existing_records.add((record.period, record.department_code))
        
        preview_items = []
        statistics = {"total": 0, "new_count": 0, "update_count": 0, "skip_count": 0, "error_count": 0}
        
        def parse_decimal(value) -> Decimal:
            if value is None:
                return Decimal("0")
            try:
                s = str(value).strip().replace(",", "")
                if not s or s == "-":
                    return Decimal("0")
                return Decimal(s)
            except:
                return Decimal("0")
        
        def parse_period(value) -> str:
            if value is None:
                return ""
            s = str(value).strip()
            # 尝试解析各种格式
            if re.match(r'^\d{4}-\d{2}$', s):
                return s
            if re.match(r'^\d{4}/\d{2}$', s):
                return s.replace("/", "-")
            if re.match(r'^\d{6}$', s):
                return f"{s[:4]}-{s[4:]}"
            if re.match(r'^\d{4}\.\d{2}$', s):
                return s.replace(".", "-")
            return s
        
        for row in data_rows:
            # 获取年月
            period = ""
            if period_col >= 0 and period_col < len(row):
                period = parse_period(row[period_col])
            
            if not period:
                continue
            
            # 获取科室信息
            excel_dept_name = ""
            dept_code = ""
            dept_name = ""
            
            if match_by == "code":
                if dept_code_col >= 0 and dept_code_col < len(row):
                    dept_code = str(row[dept_code_col]).strip() if row[dept_code_col] else ""
                if dept_name_col >= 0 and dept_name_col < len(row):
                    excel_dept_name = str(row[dept_name_col]).strip() if row[dept_name_col] else ""
                    dept_name = excel_dept_name
                
                if dept_code:
                    dept = all_departments_by_code.get(dept_code)
                    if dept:
                        dept_name = dept.accounting_unit_name or dept.his_name
            else:
                if dept_name_col >= 0 and dept_name_col < len(row):
                    excel_dept_name = str(row[dept_name_col]).strip() if row[dept_name_col] else ""
                    excel_dept_name = re.sub(r'[\t\r\n]+', '', excel_dept_name)
                    excel_dept_name = re.sub(r'\s+', ' ', excel_dept_name).strip()
                
                if excel_dept_name:
                    dept_code = name_to_code.get(excel_dept_name, "")
                    if dept_code:
                        dept = all_departments_by_code.get(dept_code)
                        if dept:
                            dept_name = dept.accounting_unit_name or dept.his_name
                        else:
                            dept_name = excel_dept_name
                    else:
                        dept_name = excel_dept_name
            
            if not dept_code and not dept_name:
                continue
            
            # 获取费用数据
            personnel_cost = parse_decimal(row[personnel_col] if personnel_col >= 0 and personnel_col < len(row) else None)
            material_cost = parse_decimal(row[material_col] if material_col >= 0 and material_col < len(row) else None)
            medicine_cost = parse_decimal(row[medicine_col] if medicine_col >= 0 and medicine_col < len(row) else None)
            depreciation_cost = parse_decimal(row[depreciation_col] if depreciation_col >= 0 and depreciation_col < len(row) else None)
            other_cost = parse_decimal(row[other_col] if other_col >= 0 and other_col < len(row) else None)
            
            statistics["total"] += 1
            
            # 检查状态
            status = "new"
            message = ""
            
            if not dept_code:
                status = "skip"
                message = "未选择匹配科室，将跳过导入"
                statistics["skip_count"] += 1
            elif (period, dept_code) in existing_records:
                status = "update"
                message = "该记录已存在，将被覆盖"
                statistics["update_count"] += 1
            else:
                statistics["new_count"] += 1
            
            preview_items.append({
                "period": period,
                "department_code": dept_code or "",
                "department_name": dept_name or "",
                "excel_department_name": excel_dept_name or "",
                "personnel_cost": personnel_cost,
                "material_cost": material_cost,
                "medicine_cost": medicine_cost,
                "depreciation_cost": depreciation_cost,
                "other_cost": other_cost,
                "status": status,
                "message": message
            })
        
        cls._sessions[session_id]["preview_items"] = preview_items
        
        return {
            "preview_items": preview_items,
            "statistics": statistics
        }
    
    @classmethod
    def execute_import(
        cls,
        session_id: str,
        confirmed_items: Optional[List[Dict[str, Any]]],
        db: Session,
        hospital_id: int
    ) -> Dict[str, Any]:
        """执行导入"""
        import logging
        logger = logging.getLogger(__name__)
        
        if confirmed_items:
            items_to_import = confirmed_items
        else:
            session_data = cls._sessions.get(session_id)
            if not session_data:
                raise ValueError("会话已过期或不存在")
            
            preview_items = session_data.get("preview_items")
            if not preview_items:
                raise ValueError("缺少预览数据")
            
            items_to_import = preview_items
        
        success_count = 0
        update_count = 0
        skip_count = 0
        error_count = 0
        errors = []
        
        def get_field(item, field_name, default=None):
            if hasattr(item, field_name):
                return getattr(item, field_name, default)
            elif isinstance(item, dict):
                return item.get(field_name, default)
            return default
        
        for item in items_to_import:
            try:
                item_status = get_field(item, "status")
                
                # 跳过用户选择不导入的记录
                if item_status == "skip":
                    skip_count += 1
                    continue
                
                if item_status == "error":
                    error_count += 1
                    errors.append({
                        "period": get_field(item, "period", ""),
                        "department_code": get_field(item, "department_code", ""),
                        "reason": get_field(item, "message", "未知错误")
                    })
                    continue
                
                period = get_field(item, "period")
                dept_code = get_field(item, "department_code")
                
                if not period or not dept_code:
                    error_count += 1
                    errors.append({
                        "period": period or "",
                        "department_code": dept_code or "",
                        "reason": "缺少必要字段"
                    })
                    continue
                
                existing = db.query(CostReport).filter(
                    CostReport.hospital_id == hospital_id,
                    CostReport.period == period,
                    CostReport.department_code == dept_code
                ).first()
                
                def to_decimal(val):
                    if val is None:
                        return Decimal("0")
                    if isinstance(val, Decimal):
                        return val
                    try:
                        return Decimal(str(val))
                    except:
                        return Decimal("0")
                
                if existing:
                    existing.department_name = get_field(item, "department_name", existing.department_name)
                    existing.personnel_cost = to_decimal(get_field(item, "personnel_cost"))
                    existing.material_cost = to_decimal(get_field(item, "material_cost"))
                    existing.medicine_cost = to_decimal(get_field(item, "medicine_cost"))
                    existing.depreciation_cost = to_decimal(get_field(item, "depreciation_cost"))
                    existing.other_cost = to_decimal(get_field(item, "other_cost"))
                    update_count += 1
                else:
                    record = CostReport(
                        hospital_id=hospital_id,
                        period=period,
                        department_code=dept_code,
                        department_name=get_field(item, "department_name", ""),
                        personnel_cost=to_decimal(get_field(item, "personnel_cost")),
                        material_cost=to_decimal(get_field(item, "material_cost")),
                        medicine_cost=to_decimal(get_field(item, "medicine_cost")),
                        depreciation_cost=to_decimal(get_field(item, "depreciation_cost")),
                        other_cost=to_decimal(get_field(item, "other_cost"))
                    )
                    db.add(record)
                    success_count += 1
                
                db.flush()
                
            except Exception as e:
                db.rollback()
                error_count += 1
                errors.append({
                    "period": get_field(item, "period", ""),
                    "department_code": get_field(item, "department_code", ""),
                    "reason": str(e)
                })
        
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            raise ValueError(f"提交事务失败: {str(e)}")
        
        if session_id and session_id in cls._sessions:
            del cls._sessions[session_id]
        
        return {
            "success": True,
            "report": {
                "success_count": success_count,
                "update_count": update_count,
                "skip_count": skip_count,
                "error_count": error_count,
                "errors": errors[:100]
            }
        }
