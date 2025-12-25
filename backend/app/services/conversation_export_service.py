"""
智能问数系统 - 对话消息导出服务

支持导出格式：
- Markdown: 适用于指标口径查询结果
- PDF: 适用于指标口径查询结果
- Excel: 适用于数据智能查询结果（表格数据）
- CSV: 适用于数据智能查询结果（表格数据）

需求 12.1: 当用户导出指标口径结果时，智能数据问答模块应生成带有格式化表格的Markdown文件
需求 12.2: 当用户将指标口径结果导出为PDF时，智能数据问答模块应生成格式正确的PDF文档
需求 12.3: 当用户将查询数据导出为Excel时，智能数据问答模块应生成包含数据和列标题的Excel文件
需求 12.4: 当用户将查询数据导出为CSV时，智能数据问答模块应生成UTF-8编码的CSV文件
"""
import csv
import html
import logging
from io import BytesIO, StringIO
from typing import Dict, Any, List, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

logger = logging.getLogger(__name__)


class ConversationExportService:
    """对话消息导出服务"""
    
    # 注册中文字体
    _font_registered = False
    
    @classmethod
    def _register_fonts(cls):
        """注册中文字体"""
        if cls._font_registered:
            return
        
        try:
            # Windows 宋体
            pdfmetrics.registerFont(TTFont('SimSun', 'C:/Windows/Fonts/simsun.ttc'))
            cls._font_registered = True
            logger.info("已注册中文字体: SimSun")
        except Exception as e:
            logger.warning(f"注册中文字体失败: {e}，将使用默认字体")
    
    @staticmethod
    def export_to_markdown(
        content: str,
        content_type: str,
        metadata: Optional[Dict[str, Any]] = None,
        title: Optional[str] = None
    ) -> BytesIO:
        """
        导出为Markdown格式
        
        Args:
            content: 消息内容
            content_type: 内容类型 (text, table, code)
            metadata: 元数据（表格数据、代码语言等）
            title: 文档标题
            
        Returns:
            BytesIO: Markdown文件的字节流
        """
        lines = []
        
        # 标题
        if title:
            lines.append(f"# {title}")
            lines.append("")
        
        # 导出时间
        lines.append(f"> 导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        
        if content_type == "table" and metadata:
            # 表格内容
            columns = metadata.get("columns", [])
            rows = metadata.get("rows", [])
            
            if columns and rows:
                # 表头
                lines.append("| " + " | ".join(str(col) for col in columns) + " |")
                # 分隔线
                lines.append("| " + " | ".join("---" for _ in columns) + " |")
                # 数据行
                for row in rows:
                    if isinstance(row, list):
                        lines.append("| " + " | ".join(str(cell) if cell is not None else "" for cell in row) + " |")
                    elif isinstance(row, dict):
                        cells = [str(row.get(col, "")) for col in columns]
                        lines.append("| " + " | ".join(cells) + " |")
                
                lines.append("")
                
                # 统计信息
                total_rows = metadata.get("total_rows", len(rows))
                lines.append(f"*共 {total_rows} 条记录*")
            else:
                lines.append(content)
        
        elif content_type == "code" and metadata:
            # 代码块
            language = metadata.get("language", "sql")
            code = metadata.get("code", content)
            lines.append(f"```{language}")
            lines.append(code)
            lines.append("```")
        
        else:
            # 普通文本
            lines.append(content)
        
        # 生成文件
        output = BytesIO()
        output.write("\n".join(lines).encode("utf-8"))
        output.seek(0)
        
        return output
    
    @classmethod
    def export_to_pdf(
        cls,
        content: str,
        content_type: str,
        metadata: Optional[Dict[str, Any]] = None,
        title: Optional[str] = None
    ) -> BytesIO:
        """
        导出为PDF格式，支持Markdown渲染
        
        Args:
            content: 消息内容（可能包含Markdown格式）
            content_type: 内容类型 (text, table)
            metadata: 元数据（表格数据等）
            title: 文档标题
            
        Returns:
            BytesIO: PDF文件的字节流
        """
        cls._register_fonts()
        
        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )
        
        # 样式
        styles = getSampleStyleSheet()
        
        # 尝试使用中文字体
        font_name = 'SimSun' if cls._font_registered else 'Helvetica'
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=16,
            spaceAfter=12,
            alignment=1  # 居中
        )
        
        h1_style = ParagraphStyle(
            'CustomH1',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=14,
            spaceBefore=12,
            spaceAfter=8,
        )
        
        h2_style = ParagraphStyle(
            'CustomH2',
            parent=styles['Heading2'],
            fontName=font_name,
            fontSize=12,
            spaceBefore=10,
            spaceAfter=6,
        )
        
        h3_style = ParagraphStyle(
            'CustomH3',
            parent=styles['Heading3'],
            fontName=font_name,
            fontSize=11,
            spaceBefore=8,
            spaceAfter=4,
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            leading=14,
            spaceAfter=6
        )
        
        info_style = ParagraphStyle(
            'CustomInfo',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=12
        )
        
        code_style = ParagraphStyle(
            'CustomCode',
            parent=styles['Normal'],
            fontName='Courier',
            fontSize=9,
            leading=12,
            backColor=colors.HexColor('#F5F5F5'),
            leftIndent=10,
            rightIndent=10,
            spaceBefore=6,
            spaceAfter=6,
        )
        
        list_style = ParagraphStyle(
            'CustomList',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            leading=14,
            leftIndent=20,
            spaceAfter=4,
        )
        
        bold_style = ParagraphStyle(
            'CustomBold',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            leading=14,
            spaceAfter=6,
        )
        
        elements = []
        
        # 标题
        if title:
            elements.append(Paragraph(html.escape(title), title_style))
        
        # 导出时间
        export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elements.append(Paragraph(f"导出时间: {export_time}", info_style))
        elements.append(Spacer(1, 10))
        
        if content_type == "table" and metadata:
            # 表格内容（从metadata构建）
            columns = metadata.get("columns", [])
            rows = metadata.get("rows", [])
            
            if columns and rows:
                table_element = cls._build_pdf_table(columns, rows, font_name, normal_style)
                elements.append(table_element)
                elements.append(Spacer(1, 10))
                
                # 统计信息
                total_rows = metadata.get("total_rows", len(rows))
                elements.append(Paragraph(f"共 {total_rows} 条记录", info_style))
            else:
                elements.append(Paragraph(html.escape(content), normal_style))
        
        else:
            # 解析Markdown内容
            md_elements = cls._parse_markdown_to_elements(
                content, font_name, 
                h1_style, h2_style, h3_style, 
                normal_style, code_style, list_style, bold_style
            )
            elements.extend(md_elements)
        
        # 生成PDF
        doc.build(elements)
        output.seek(0)
        
        return output
    
    @classmethod
    def _build_pdf_table(
        cls,
        columns: List[str],
        rows: List,
        font_name: str,
        normal_style: ParagraphStyle
    ):
        """构建PDF表格"""
        table_data = []
        
        # 表头
        header_row = [Paragraph(html.escape(str(col)), normal_style) for col in columns]
        table_data.append(header_row)
        
        # 数据行
        for row in rows:
            if isinstance(row, list):
                data_row = [
                    Paragraph(html.escape(str(cell)) if cell is not None else "", normal_style)
                    for cell in row
                ]
            elif isinstance(row, dict):
                data_row = [
                    Paragraph(html.escape(str(row.get(col, ""))), normal_style)
                    for col in columns
                ]
            else:
                data_row = [Paragraph("", normal_style)] * len(columns)
            table_data.append(data_row)
        
        # 计算列宽
        available_width = A4[0] - 40*mm
        col_width = available_width / len(columns)
        col_widths = [col_width] * len(columns)
        
        # 创建表格
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        
        return table
    
    @classmethod
    def _parse_markdown_to_elements(
        cls,
        content: str,
        font_name: str,
        h1_style: ParagraphStyle,
        h2_style: ParagraphStyle,
        h3_style: ParagraphStyle,
        normal_style: ParagraphStyle,
        code_style: ParagraphStyle,
        list_style: ParagraphStyle,
        bold_style: ParagraphStyle,
    ) -> List:
        """
        解析Markdown内容为PDF元素
        
        支持：
        - 标题 (# ## ###)
        - 列表 (- *)
        - 代码块 (```)
        - 粗体 (**text**)
        - 表格 (| col1 | col2 |)
        """
        import re
        
        # 预处理：处理HTML实体（但不处理<br>，稍后按上下文处理）
        # 处理 &nbsp; 转换为空格
        content = content.replace('&nbsp;', ' ')
        # 处理 &lt; &gt; &amp; 等HTML实体
        content = content.replace('&lt;', '<')
        content = content.replace('&gt;', '>')
        content = content.replace('&amp;', '&')
        content = content.replace('&quot;', '"')
        content = content.replace('&#39;', "'")
        
        # 智能处理 <br> 标签：
        # 1. 在表格行内（包含 | 的行），<br> 转换为空格
        # 2. 在其他地方，<br> 转换为换行符
        processed_lines = []
        for line in content.split('\n'):
            if '|' in line and re.search(r'\|.*\|', line):
                # 表格行：<br> 转换为空格
                line = re.sub(r'<br\s*/?>', ' ', line, flags=re.IGNORECASE)
            else:
                # 非表格行：<br> 转换为换行符，然后拆分
                if re.search(r'<br\s*/?>', line, flags=re.IGNORECASE):
                    sub_lines = re.split(r'<br\s*/?>', line, flags=re.IGNORECASE)
                    processed_lines.extend(sub_lines)
                    continue
            processed_lines.append(line)
        
        lines = processed_lines
        elements = []
        
        i = 0
        in_code_block = False
        code_lines = []
        code_language = ""
        
        # 用于收集表格行
        table_lines = []
        
        while i < len(lines):
            line = lines[i]
            
            # 代码块处理
            if line.strip().startswith('```'):
                if not in_code_block:
                    # 开始代码块
                    in_code_block = True
                    code_language = line.strip()[3:].strip()
                    code_lines = []
                else:
                    # 结束代码块
                    in_code_block = False
                    if code_lines:
                        code_text = '\n'.join(code_lines)
                        # 转义HTML特殊字符
                        code_text = html.escape(code_text)
                        # 保留换行和空格
                        code_text = code_text.replace('\n', '<br/>')
                        code_text = code_text.replace(' ', '&nbsp;')
                        elements.append(Paragraph(code_text, code_style))
                        elements.append(Spacer(1, 6))
                i += 1
                continue
            
            if in_code_block:
                code_lines.append(line)
                i += 1
                continue
            
            # 表格处理 - 检测以 | 开头和结尾的行
            stripped_line = line.strip()
            if stripped_line.startswith('|') and stripped_line.endswith('|'):
                table_lines.append(line)
                i += 1
                continue
            elif table_lines:
                # 表格结束，渲染表格
                table_element = cls._parse_markdown_table(table_lines, font_name, normal_style)
                if table_element:
                    elements.append(table_element)
                    elements.append(Spacer(1, 10))
                table_lines = []
            
            # 标题处理
            if line.startswith('### '):
                text = line[4:].strip()
                text = cls._process_inline_markdown(text)
                elements.append(Paragraph(text, h3_style))
                i += 1
                continue
            
            if line.startswith('## '):
                text = line[3:].strip()
                text = cls._process_inline_markdown(text)
                elements.append(Paragraph(text, h2_style))
                i += 1
                continue
            
            if line.startswith('# '):
                text = line[2:].strip()
                text = cls._process_inline_markdown(text)
                elements.append(Paragraph(text, h1_style))
                i += 1
                continue
            
            # 列表处理
            list_match = re.match(r'^(\s*)[-*]\s+(.+)$', line)
            if list_match:
                indent = len(list_match.group(1))
                text = list_match.group(2)
                text = cls._process_inline_markdown(text)
                # 添加列表符号
                bullet = "• " if indent == 0 else "  ◦ "
                elements.append(Paragraph(bullet + text, list_style))
                i += 1
                continue
            
            # 引用处理
            if line.startswith('> '):
                text = line[2:].strip()
                text = cls._process_inline_markdown(text)
                quote_style = ParagraphStyle(
                    'Quote',
                    parent=normal_style,
                    leftIndent=20,
                    textColor=colors.HexColor('#666666'),
                    borderColor=colors.HexColor('#CCCCCC'),
                    borderWidth=1,
                    borderPadding=5,
                )
                elements.append(Paragraph(text, quote_style))
                i += 1
                continue
            
            # 空行
            if not line.strip():
                elements.append(Spacer(1, 6))
                i += 1
                continue
            
            # 普通段落
            text = cls._process_inline_markdown(line)
            elements.append(Paragraph(text, normal_style))
            i += 1
        
        # 处理剩余的表格
        if table_lines:
            table_element = cls._parse_markdown_table(table_lines, font_name, normal_style)
            if table_element:
                elements.append(table_element)
        
        return elements
    
    @classmethod
    def _process_inline_markdown(cls, text: str) -> str:
        """
        处理行内Markdown格式
        
        支持：
        - **粗体**
        - *斜体*
        - `代码`
        """
        import re
        
        # 先转义HTML
        text = html.escape(text)
        
        # 粗体 **text** -> <b>text</b>
        text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text)
        
        # 斜体 *text* -> <i>text</i>
        text = re.sub(r'\*(.+?)\*', r'<i>\1</i>', text)
        
        # 行内代码 `code` -> <font face="Courier">code</font>
        text = re.sub(r'`(.+?)`', r'<font face="Courier" color="#C7254E">\1</font>', text)
        
        return text
    
    @classmethod
    def _parse_markdown_table(
        cls,
        table_lines: List[str],
        font_name: str,
        normal_style: ParagraphStyle
    ):
        """
        解析Markdown表格为PDF Table
        
        Args:
            table_lines: 表格行列表
            font_name: 字体名称
            normal_style: 普通文本样式
            
        Returns:
            Table对象或None
        """
        if len(table_lines) < 1:
            return None
        
        # 解析表格行
        rows = []
        header_row_index = -1
        
        for idx, line in enumerate(table_lines):
            # 去除首尾空白
            line = line.strip()
            
            # 去除首尾的 | (如果有)
            if line.startswith('|'):
                line = line[1:]
            if line.endswith('|'):
                line = line[:-1]
            
            # 分割单元格
            cells = [cell.strip() for cell in line.split('|')]
            
            # 检查是否是分隔行 (|---|---|) 或 (|:---|:---:|)
            is_separator = all(
                set(cell.strip().replace(':', '')) <= {'-'} and len(cell.strip()) > 0
                for cell in cells
            )
            
            if is_separator:
                # 记录分隔行位置，它上面的是表头
                header_row_index = len(rows) - 1 if rows else -1
                continue
            
            # 跳过空行
            if not any(cell for cell in cells):
                continue
            
            rows.append(cells)
        
        if not rows:
            return None
        
        # 确保所有行有相同的列数
        max_cols = max(len(row) for row in rows)
        for row in rows:
            while len(row) < max_cols:
                row.append('')
        
        # 构建表格数据
        table_data = []
        for row in rows:
            data_row = [
                Paragraph(cls._process_inline_markdown(cell), normal_style)
                for cell in row
            ]
            table_data.append(data_row)
        
        if not table_data:
            return None
        
        # 计算列宽 - 根据内容自适应
        num_cols = max(len(row) for row in table_data)
        available_width = A4[0] - 40*mm
        
        # 计算每列的最大内容长度
        col_lengths = [0] * num_cols
        for row in rows:
            for col_idx, cell in enumerate(row):
                if col_idx < num_cols:
                    # 中文字符算2个单位
                    cell_len = sum(2 if ord(c) > 127 else 1 for c in cell)
                    col_lengths[col_idx] = max(col_lengths[col_idx], cell_len)
        
        # 根据内容长度分配列宽
        total_length = sum(col_lengths) or 1
        col_widths = [
            max(available_width * (length / total_length), 30*mm)
            for length in col_lengths
        ]
        
        # 如果总宽度超出，按比例缩放
        total_width = sum(col_widths)
        if total_width > available_width:
            scale = available_width / total_width
            col_widths = [w * scale for w in col_widths]
        
        # 创建表格
        table = Table(table_data, colWidths=col_widths)
        
        # 确定表头行数（默认第一行是表头）
        header_rows = 1 if header_row_index >= 0 or len(rows) > 1 else 0
        
        style_commands = [
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]
        
        # 如果有表头，添加表头样式
        if header_rows > 0 and len(table_data) > 0:
            style_commands.extend([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
            ])
            # 数据行交替背景色
            if len(table_data) > 1:
                style_commands.append(
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
                )
        
        table.setStyle(TableStyle(style_commands))
        
        return table

    
    @staticmethod
    def export_to_excel(
        content: str,
        content_type: str,
        metadata: Optional[Dict[str, Any]] = None,
        title: Optional[str] = None
    ) -> BytesIO:
        """
        导出为Excel格式
        
        Args:
            content: 消息内容
            content_type: 内容类型 (table)
            metadata: 元数据（表格数据）
            title: 文档标题
            
        Returns:
            BytesIO: Excel文件的字节流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "查询结果"
        
        # 样式定义
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        data_alignment_right = Alignment(horizontal='right', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_num = 1
        
        # 标题行
        if title:
            title_cell = ws.cell(row=row_num, column=1, value=title)
            title_cell.font = Font(name='微软雅黑', size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num += 1
        
        # 导出时间
        time_cell = ws.cell(row=row_num, column=1, value=f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        time_cell.font = Font(name='微软雅黑', size=9, color='808080')
        row_num += 2
        
        if content_type == "table" and metadata:
            columns = metadata.get("columns", [])
            rows = metadata.get("rows", [])
            
            if columns:
                # 合并标题单元格
                if title and len(columns) > 1:
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
                
                # 表头
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                ws.row_dimensions[row_num].height = 25
                row_num += 1
                
                # 数据行
                for row_data in rows:
                    if isinstance(row_data, list):
                        for col_idx, cell_value in enumerate(row_data, start=1):
                            cell = ws.cell(row=row_num, column=col_idx, value=cell_value)
                            cell.font = data_font
                            cell.border = border
                            # 数字右对齐
                            if isinstance(cell_value, (int, float)):
                                cell.alignment = data_alignment_right
                                if isinstance(cell_value, float):
                                    cell.number_format = '#,##0.00'
                            else:
                                cell.alignment = data_alignment
                    elif isinstance(row_data, dict):
                        for col_idx, col_name in enumerate(columns, start=1):
                            cell_value = row_data.get(col_name, "")
                            cell = ws.cell(row=row_num, column=col_idx, value=cell_value)
                            cell.font = data_font
                            cell.border = border
                            if isinstance(cell_value, (int, float)):
                                cell.alignment = data_alignment_right
                                if isinstance(cell_value, float):
                                    cell.number_format = '#,##0.00'
                            else:
                                cell.alignment = data_alignment
                    
                    ws.row_dimensions[row_num].height = 20
                    row_num += 1
                
                # 自动调整列宽
                for col_idx, col_name in enumerate(columns, start=1):
                    # 计算列宽（基于列名长度和数据）
                    max_length = len(str(col_name))
                    for row_data in rows[:100]:  # 只检查前100行
                        if isinstance(row_data, list) and col_idx <= len(row_data):
                            cell_value = row_data[col_idx - 1]
                        elif isinstance(row_data, dict):
                            cell_value = row_data.get(col_name, "")
                        else:
                            cell_value = ""
                        max_length = max(max_length, len(str(cell_value)) if cell_value else 0)
                    
                    # 设置列宽（中文字符约占2个单位）
                    adjusted_width = min(max(max_length * 1.2 + 2, 10), 50)
                    ws.column_dimensions[chr(64 + col_idx)].width = adjusted_width
        else:
            # 非表格内容，直接写入文本
            ws.cell(row=row_num, column=1, value=content)
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_to_csv(
        content: str,
        content_type: str,
        metadata: Optional[Dict[str, Any]] = None
    ) -> BytesIO:
        """
        导出为CSV格式（UTF-8编码，带BOM）
        
        Args:
            content: 消息内容
            content_type: 内容类型 (table)
            metadata: 元数据（表格数据）
            
        Returns:
            BytesIO: CSV文件的字节流
        """
        output = BytesIO()
        
        # 写入UTF-8 BOM，确保Excel正确识别编码
        output.write(b'\xef\xbb\xbf')
        
        if content_type == "table" and metadata:
            columns = metadata.get("columns", [])
            rows = metadata.get("rows", [])
            
            if columns:
                # 使用StringIO写入CSV
                string_buffer = StringIO()
                writer = csv.writer(string_buffer, quoting=csv.QUOTE_MINIMAL)
                
                # 写入表头
                writer.writerow(columns)
                
                # 写入数据行
                for row_data in rows:
                    if isinstance(row_data, list):
                        writer.writerow(row_data)
                    elif isinstance(row_data, dict):
                        writer.writerow([row_data.get(col, "") for col in columns])
                
                # 转换为字节
                output.write(string_buffer.getvalue().encode('utf-8'))
            else:
                output.write(content.encode('utf-8'))
        else:
            output.write(content.encode('utf-8'))
        
        output.seek(0)
        return output
    
    @classmethod
    def export_message(
        cls,
        content: str,
        content_type: str,
        export_format: str,
        metadata: Optional[Dict[str, Any]] = None,
        title: Optional[str] = None
    ) -> tuple[BytesIO, str, str]:
        """
        导出消息内容
        
        Args:
            content: 消息内容
            content_type: 内容类型
            export_format: 导出格式 (markdown, pdf, excel, csv)
            metadata: 元数据
            title: 文档标题
            
        Returns:
            tuple: (文件字节流, 文件扩展名, MIME类型)
            
        Raises:
            ValueError: 不支持的导出格式
        """
        if export_format == "markdown":
            file_data = cls.export_to_markdown(content, content_type, metadata, title)
            return file_data, ".md", "text/markdown; charset=utf-8"
        
        elif export_format == "pdf":
            file_data = cls.export_to_pdf(content, content_type, metadata, title)
            return file_data, ".pdf", "application/pdf"
        
        elif export_format == "excel":
            file_data = cls.export_to_excel(content, content_type, metadata, title)
            return file_data, ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        elif export_format == "csv":
            file_data = cls.export_to_csv(content, content_type, metadata)
            return file_data, ".csv", "text/csv; charset=utf-8"
        
        else:
            raise ValueError(f"不支持的导出格式: {export_format}")
