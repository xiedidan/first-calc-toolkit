"""
报表导出服务
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict, Any
from decimal import Decimal
import zipfile


class ExportService:
    """报表导出服务"""
    
    @staticmethod
    def _format_version(version: str = None) -> str:
        """格式化版本号，用于文件名"""
        if not version:
            return ""
        # 版本号本身可能已经包含v前缀，不重复添加
        v = version.strip()
        return f"_{v}"
    
    @staticmethod
    def export_summary_to_excel(summary_data: dict, period: str, hospital_name: str = None, version: str = None) -> BytesIO:
        """
        导出汇总表到Excel
        
        Args:
            summary_data: 汇总数据，包含summary和departments
            period: 评估月份
            
        Returns:
            BytesIO: Excel文件的字节流
        """
        wb = Workbook()
        ws = wb.active
        # Sheet名称不加版本号
        ws.title = "汇总表"
        
        # 设置列宽（与页面列一致）
        ws.column_dimensions['A'].width = 15  # 科室代码
        ws.column_dimensions['B'].width = 22  # 科室名称
        ws.column_dimensions['C'].width = 16.5  # 医生价值
        ws.column_dimensions['D'].width = 13  # 医生占比
        ws.column_dimensions['E'].width = 16.5  # 护理价值
        ws.column_dimensions['F'].width = 13  # 护理占比
        ws.column_dimensions['G'].width = 16.5  # 医技价值
        ws.column_dimensions['H'].width = 13  # 医技占比
        ws.column_dimensions['I'].width = 18  # 科室总价值
        ws.column_dimensions['J'].width = 18  # 参考总价值
        ws.column_dimensions['K'].width = 14  # 核算/实发
        ws.column_dimensions['L'].width = 16  # 环期价值
        ws.column_dimensions['M'].width = 14  # 当期/环期
        ws.column_dimensions['N'].width = 16  # 同期价值
        ws.column_dimensions['O'].width = 14  # 当期/同期
        
        # 标题行（表头加版本号）
        version_suffix = ExportService._format_version(version)
        title_cell = ws.cell(row=1, column=1, value=f"科室业务价值汇总（{period}）{version_suffix}")
        title_cell.font = Font(name='微软雅黑', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:O1')
        ws.row_dimensions[1].height = 30
        
        # 表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 第一层表头（序列分组）
        row = 2
        ws.merge_cells(f'A{row}:A{row+1}')  # 科室代码
        ws.cell(row=row, column=1, value='科室代码')
        
        ws.merge_cells(f'B{row}:B{row+1}')  # 科室名称
        ws.cell(row=row, column=2, value='科室名称')
        
        ws.merge_cells(f'C{row}:D{row}')  # 医生序列
        ws.cell(row=row, column=3, value='医生序列')
        
        ws.merge_cells(f'E{row}:F{row}')  # 护理序列
        ws.cell(row=row, column=5, value='护理序列')
        
        ws.merge_cells(f'G{row}:H{row}')  # 医技序列
        ws.cell(row=row, column=7, value='医技序列')
        
        ws.merge_cells(f'I{row}:I{row+1}')  # 科室总价值
        ws.cell(row=row, column=9, value='科室总价值')
        
        ws.merge_cells(f'J{row}:J{row+1}')  # 参考总价值
        ws.cell(row=row, column=10, value='参考总价值')
        
        ws.merge_cells(f'K{row}:K{row+1}')  # 核算/实发
        ws.cell(row=row, column=11, value='核算/实发')
        
        ws.merge_cells(f'L{row}:L{row+1}')  # 环期价值
        ws.cell(row=row, column=12, value='环期价值')
        
        ws.merge_cells(f'M{row}:M{row+1}')  # 当期/环期
        ws.cell(row=row, column=13, value='当期/环期')
        
        ws.merge_cells(f'N{row}:N{row+1}')  # 同期价值
        ws.cell(row=row, column=14, value='同期价值')
        
        ws.merge_cells(f'O{row}:O{row+1}')  # 当期/同期
        ws.cell(row=row, column=15, value='当期/同期')
        
        # 第二层表头（价值/占比）
        row = 3
        ws.cell(row=row, column=3, value='价值')
        ws.cell(row=row, column=4, value='占比')
        ws.cell(row=row, column=5, value='价值')
        ws.cell(row=row, column=6, value='占比')
        ws.cell(row=row, column=7, value='价值')
        ws.cell(row=row, column=8, value='占比')
        
        # 应用表头样式
        for r in [2, 3]:
            for c in range(1, 16):
                cell = ws.cell(row=r, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
        
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 25
        
        # 数据行样式
        data_alignment_left = Alignment(horizontal='left', vertical='center')
        data_alignment_right = Alignment(horizontal='right', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')
        
        # 全院汇总样式（高亮）
        summary_font = Font(name='微软雅黑', size=11, bold=True)
        summary_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        # 普通数据样式
        normal_font = Font(name='微软雅黑', size=10)
        
        def format_ratio_value(value):
            """格式化比值（大于1或小于1的比例）"""
            if value is None:
                return None
            return float(value)
        
        def write_data_row(ws, row, data, font, fill=None):
            """写入数据行"""
            ws.cell(row=row, column=1, value=data.get('department_code') or '')
            ws.cell(row=row, column=2, value=data['department_name'])
            ws.cell(row=row, column=3, value=float(data['doctor_value']))
            ws.cell(row=row, column=4, value=float(data['doctor_ratio']) / 100)
            ws.cell(row=row, column=5, value=float(data['nurse_value']))
            ws.cell(row=row, column=6, value=float(data['nurse_ratio']) / 100)
            ws.cell(row=row, column=7, value=float(data['tech_value']))
            ws.cell(row=row, column=8, value=float(data['tech_ratio']) / 100)
            ws.cell(row=row, column=9, value=float(data['total_value']))
            
            # 参考总价值
            ref_value = data.get('reference_value')
            ws.cell(row=row, column=10, value=float(ref_value) if ref_value is not None else None)
            
            # 核算/实发比例
            actual_ref_ratio = data.get('actual_reference_ratio')
            ws.cell(row=row, column=11, value=format_ratio_value(actual_ref_ratio))
            
            # 环期价值
            mom_value = data.get('mom_value')
            ws.cell(row=row, column=12, value=float(mom_value) if mom_value is not None else None)
            
            # 当期/环期比例
            mom_ratio = data.get('mom_ratio')
            ws.cell(row=row, column=13, value=format_ratio_value(mom_ratio))
            
            # 同期价值
            yoy_value = data.get('yoy_value')
            ws.cell(row=row, column=14, value=float(yoy_value) if yoy_value is not None else None)
            
            # 当期/同期比例
            yoy_ratio = data.get('yoy_ratio')
            ws.cell(row=row, column=15, value=format_ratio_value(yoy_ratio))
            
            # 应用样式
            for c in range(1, 16):
                cell = ws.cell(row=row, column=c)
                cell.font = font
                if fill:
                    cell.fill = fill
                cell.border = border
                
                if c in [1, 2]:
                    cell.alignment = data_alignment_left
                elif c in [4, 6, 8]:  # 序列占比列
                    cell.alignment = data_alignment_center
                    cell.number_format = '0.00%'
                elif c in [11, 13, 15]:  # 比值列（核算/实发、当期/环期、当期/同期）
                    cell.alignment = data_alignment_center
                    if cell.value is not None:
                        cell.number_format = '0.00%'
                else:  # 价值列
                    cell.alignment = data_alignment_right
                    if cell.value is not None:
                        cell.number_format = '#,##0.00'
        
        # 写入数据
        row = 4
        
        # 全院汇总行
        summary = summary_data['summary']
        write_data_row(ws, row, summary, summary_font, summary_fill)
        ws.row_dimensions[row].height = 22
        row += 1
        
        # 各科室数据行
        for dept in summary_data['departments']:
            write_data_row(ws, row, dept, normal_font)
            ws.row_dimensions[row].height = 20
            row += 1
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_detail_to_excel(dept_name: str, period: str, detail_data: Dict[str, List[Dict]], hospital_name: str = None, version: str = None) -> BytesIO:
        """
        导出单个科室的明细表到Excel
        
        Args:
            dept_name: 科室名称
            period: 评估月份
            detail_data: 明细数据，包含doctor、nurse、tech三个序列的树形数据
            
        Returns:
            BytesIO: Excel文件的字节流
        """
        wb = Workbook()
        
        # 样式定义
        title_font = Font(name='微软雅黑', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='微软雅黑', size=10)
        data_alignment_left = Alignment(horizontal='left', vertical='center')
        data_alignment_right = Alignment(horizontal='right', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 为每个序列创建一个Sheet
        sequence_names = {
            'doctor': '医生序列',
            'nurse': '护理序列',
            'tech': '医技序列'
        }
        
        first_sheet = True
        for seq_key, seq_name in sequence_names.items():
            if seq_key not in detail_data or not detail_data[seq_key]:
                continue
            
            # Sheet名称不加版本号
            sheet_title = seq_name
            
            if first_sheet:
                ws = wb.active
                ws.title = sheet_title
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_title)
            
            # 设置列宽
            ws.column_dimensions['A'].width = 33  # 维度名称
            ws.column_dimensions['B'].width = 13  # 工作量
            ws.column_dimensions['C'].width = 16  # 全院业务价值
            ws.column_dimensions['D'].width = 27  # 业务导向（增加50%：18 * 1.5）
            ws.column_dimensions['E'].width = 16  # 科室业务价值
            ws.column_dimensions['F'].width = 16  # 业务价值金额
            
            # 标题行（表头加版本号）
            version_suffix = ExportService._format_version(version)
            title_text = f"{dept_name} - {seq_name}业务价值明细（{period}）{version_suffix}"
            title_cell = ws.cell(row=1, column=1, value=title_text)
            title_cell.font = title_font
            title_cell.alignment = title_alignment
            ws.merge_cells('A1:F1')  # 改为F1（去掉占比列）
            ws.row_dimensions[1].height = 30
            
            # 表头（去掉占比列）
            headers = ['维度名称（业务价值占比）', '工作量', '全院业务价值', '业务导向', '科室业务价值', '业务价值金额']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            ws.row_dimensions[2].height = 25
            
            # 写入树形数据
            row = 3
            
            def write_tree_node(node: Dict, level: int, current_row: int) -> int:
                """递归写入树形节点"""
                # 计算缩进（每级缩进4个空格，让层级更明显）
                indent = '    ' * level
                
                # 维度名称（带占比）
                dim_name = indent + node['dimension_name']
                if node.get('ratio') is not None and node['ratio'] != 0:
                    dim_name += f"（{float(node['ratio']):.2f}%）"
                
                ws.cell(row=current_row, column=1, value=dim_name)
                ws.cell(row=current_row, column=2, value=float(node.get('workload', 0)) if node.get('workload') else None)
                ws.cell(row=current_row, column=3, value=node.get('hospital_value', '-'))
                ws.cell(row=current_row, column=4, value=node.get('business_guide', '-'))
                ws.cell(row=current_row, column=5, value=node.get('dept_value', '-'))
                ws.cell(row=current_row, column=6, value=float(node.get('amount', 0)) if node.get('amount') else None)
                
                # 应用样式（去掉占比列）
                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = data_font
                    cell.border = border
                    
                    if col == 1:  # 维度名称
                        cell.alignment = data_alignment_left
                    elif col in [2, 6]:  # 工作量、金额
                        cell.alignment = data_alignment_right
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                    else:  # 其他
                        cell.alignment = data_alignment_center
                
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                
                # 递归处理子节点
                if 'children' in node and node['children']:
                    for child in node['children']:
                        current_row = write_tree_node(child, level + 1, current_row)
                
                return current_row
            
            # 写入所有一级维度
            for node in detail_data[seq_key]:
                row = write_tree_node(node, 0, row)
        
        # 如果没有任何数据，删除默认Sheet
        if first_sheet:
            ws = wb.active
            ws.title = "无数据"
            ws.cell(row=1, column=1, value="该科室暂无业务价值数据")
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_all_reports_to_zip(
        period: str,
        summary_data: dict,
        departments_data: List[Dict[str, Any]], 
        hospital_name: str = None,
        hospital_detail_data: Dict[str, List[Dict]] = None,
        version: str = None
    ) -> BytesIO:
        """
        导出业务价值报表，打包成ZIP（包含汇总表、全院明细表和各科室明细表）
        
        Args:
            period: 评估月份
            summary_data: 汇总表数据，包含summary和departments
            departments_data: 所有科室的明细数据列表
                每个元素包含：dept_name, doctor, nurse, tech
            hospital_name: 医院名称
            hospital_detail_data: 全院汇总的明细数据（所有科室按维度累加），包含doctor、nurse、tech
            version: 模型版本号
            
        Returns:
            BytesIO: ZIP文件的字节流
        """
        zip_buffer = BytesIO()
        version_suffix = ExportService._format_version(version)
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 1. 首先添加汇总表
            if summary_data:
                summary_excel = ExportService.export_summary_to_excel(summary_data, period, hospital_name, version)
                if hospital_name:
                    summary_filename = f"{hospital_name}_科室业务价值汇总_{period}{version_suffix}.xlsx"
                else:
                    summary_filename = f"科室业务价值汇总_{period}{version_suffix}.xlsx"
                zip_file.writestr(summary_filename, summary_excel.getvalue())
            
            # 2. 添加全院汇总明细表（如果提供了汇总数据）
            if hospital_detail_data:
                hospital_excel = ExportService.export_hospital_detail_to_excel(period, hospital_detail_data, hospital_name, version)
                if hospital_name:
                    hospital_filename = f"{hospital_name}_全院业务价值明细_{period}{version_suffix}.xlsx"
                else:
                    hospital_filename = f"全院业务价值明细_{period}{version_suffix}.xlsx"
                zip_file.writestr(hospital_filename, hospital_excel.getvalue())
            
            # 3. 然后添加各科室明细表
            for dept_data in departments_data:
                dept_name = dept_data['dept_name']
                
                # 生成该科室的Excel文件
                detail_data = {
                    'doctor': dept_data.get('doctor', []),
                    'nurse': dept_data.get('nurse', []),
                    'tech': dept_data.get('tech', [])
                }
                
                excel_file = ExportService.export_detail_to_excel(dept_name, period, detail_data, hospital_name, version)
                
                # 添加到ZIP（医院名称_科室名称_业务价值明细_期间_版本.xlsx）
                if hospital_name:
                    filename = f"{hospital_name}_{dept_name}_业务价值明细_{period}{version_suffix}.xlsx"
                else:
                    filename = f"{dept_name}_业务价值明细_{period}{version_suffix}.xlsx"
                zip_file.writestr(filename, excel_file.getvalue())
        
        zip_buffer.seek(0)
        return zip_buffer

    @staticmethod
    def export_hospital_detail_to_excel(period: str, hospital_detail_data: Dict[str, List[Dict]], hospital_name: str = None, version: str = None) -> BytesIO:
        """
        导出全院汇总明细表到单个Excel文件（所有科室数据按维度累加汇总）
        
        Args:
            period: 评估月份
            hospital_detail_data: 全院汇总的明细数据，包含doctor、nurse、tech三个序列的树形数据
            hospital_name: 医院名称
            version: 模型版本号
            
        Returns:
            BytesIO: Excel文件的字节流
        """
        wb = Workbook()
        
        # 样式定义
        title_font = Font(name='微软雅黑', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='微软雅黑', size=10)
        data_alignment_left = Alignment(horizontal='left', vertical='center')
        data_alignment_right = Alignment(horizontal='right', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 为每个序列创建一个Sheet
        sequence_names = {
            'doctor': '医生序列',
            'nurse': '护理序列',
            'tech': '医技序列'
        }
        
        first_sheet = True
        for seq_key, seq_name in sequence_names.items():
            if seq_key not in hospital_detail_data or not hospital_detail_data[seq_key]:
                continue
            
            # Sheet名称不加版本号
            sheet_title = seq_name
            
            if first_sheet:
                ws = wb.active
                ws.title = sheet_title
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_title)
            
            # 设置列宽（与单科室明细表一致，但没有科室列）
            ws.column_dimensions['A'].width = 33  # 维度名称
            ws.column_dimensions['B'].width = 13  # 工作量
            ws.column_dimensions['C'].width = 16  # 全院业务价值
            ws.column_dimensions['D'].width = 27  # 业务导向
            ws.column_dimensions['E'].width = 16  # 业务价值金额
            
            # 标题行（表头加版本号）
            version_suffix = ExportService._format_version(version)
            title_prefix = f"{hospital_name} - " if hospital_name else ""
            title_text = f"{title_prefix}{seq_name}全院业务价值明细（{period}）{version_suffix}"
            title_cell = ws.cell(row=1, column=1, value=title_text)
            title_cell.font = title_font
            title_cell.alignment = title_alignment
            ws.merge_cells('A1:E1')
            ws.row_dimensions[1].height = 30
            
            # 表头（全院汇总不需要科室业务价值列，因为是汇总值）
            headers = ['维度名称（业务价值占比）', '工作量', '全院业务价值', '业务导向', '业务价值金额']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            ws.row_dimensions[2].height = 25
            
            # 写入数据
            row = 3
            
            def write_tree_node(node: Dict, level: int, current_row: int) -> int:
                """递归写入树形节点"""
                # 计算缩进
                indent = '    ' * level
                
                # 维度名称（带占比）
                dim_name = indent + node['dimension_name']
                if node.get('ratio') is not None and node['ratio'] != 0:
                    dim_name += f"（{float(node['ratio']):.2f}%）"
                
                ws.cell(row=current_row, column=1, value=dim_name)
                ws.cell(row=current_row, column=2, value=float(node.get('workload', 0)) if node.get('workload') else None)
                ws.cell(row=current_row, column=3, value=node.get('hospital_value', '-'))
                ws.cell(row=current_row, column=4, value=node.get('business_guide', '-'))
                ws.cell(row=current_row, column=5, value=float(node.get('amount', 0)) if node.get('amount') else None)
                
                # 应用样式
                for col in range(1, 6):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = data_font
                    cell.border = border
                    
                    if col == 1:  # 维度名称
                        cell.alignment = data_alignment_left
                    elif col in [2, 5]:  # 工作量、金额
                        cell.alignment = data_alignment_right
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                    else:  # 其他
                        cell.alignment = data_alignment_center
                
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                
                # 递归处理子节点
                if 'children' in node and node['children']:
                    for child in node['children']:
                        current_row = write_tree_node(child, level + 1, current_row)
                
                return current_row
            
            # 写入所有一级维度
            for node in hospital_detail_data[seq_key]:
                row = write_tree_node(node, 0, row)
        
        # 如果没有任何数据，删除默认Sheet
        if first_sheet:
            ws = wb.active
            ws.title = "无数据"
            ws.cell(row=1, column=1, value="暂无业务价值数据")
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
