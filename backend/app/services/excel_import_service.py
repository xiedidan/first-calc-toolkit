"""
通用Excel导入服务
"""
from typing import Dict, List, Any, Optional
from io import BytesIO
import openpyxl
from sqlalchemy.orm import Session


class ExcelImportService:
    """通用Excel导入服务"""
    
    def __init__(self, field_config: Dict[str, Any]):
        """
        初始化导入服务
        
        Args:
            field_config: 字段配置
                {
                    "fields": {
                        "item_code": {
                            "label": "项目编码",
                            "required": True,
                            "unique": True
                        },
                        ...
                    },
                    "default_mapping": {
                        "项目编码": "item_code",
                        ...
                    }
                }
        """
        self.field_config = field_config
    
    def parse_excel(self, file_content: bytes) -> Dict[str, Any]:
        """
        解析Excel文件
        
        Args:
            file_content: Excel文件内容
        
        Returns:
            {
                "headers": ["列1", "列2", ...],
                "preview_data": [[值1, 值2, ...], ...],
                "total_rows": 100,
                "suggested_mapping": {"列1": "field1", ...}
            }
        """
        # 验证文件内容
        if not file_content or len(file_content) == 0:
            raise ValueError("文件内容为空")
        
        # 读取Excel
        try:
            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
            ws = wb.active
        except Exception as e:
            error_msg = str(e)
            if "zip" in error_msg.lower():
                raise ValueError("文件格式错误，请上传有效的 Excel 文件（.xlsx 格式）")
            else:
                raise ValueError(f"无法读取 Excel 文件: {error_msg}")
        
        # 获取所有行
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel文件为空")
        
        # 第一行作为表头
        headers = [str(cell) if cell is not None else f"列{i+1}" 
                  for i, cell in enumerate(rows[0])]
        
        # 数据行（跳过表头）
        data_rows = rows[1:]
        total_rows = len(data_rows)
        
        # 预览数据（前10行）
        preview_data = []
        for row in data_rows[:10]:
            preview_data.append([str(cell) if cell is not None else "" 
                               for cell in row])
        
        # 建议的字段映射
        suggested_mapping = self._suggest_mapping(headers)
        
        return {
            "headers": headers,
            "preview_data": preview_data,
            "total_rows": total_rows,
            "suggested_mapping": suggested_mapping
        }
    
    def _suggest_mapping(self, headers: List[str]) -> Dict[str, str]:
        """
        根据表头建议字段映射
        
        Args:
            headers: Excel表头列表
        
        Returns:
            建议的映射关系
        """
        suggested = {}
        default_mapping = self.field_config.get("default_mapping", {})
        
        for header in headers:
            # 精确匹配
            if header in default_mapping:
                suggested[header] = default_mapping[header]
            else:
                # 模糊匹配（转小写后匹配）
                header_lower = header.lower()
                for key, value in default_mapping.items():
                    if key.lower() == header_lower:
                        suggested[header] = value
                        break
        
        return suggested
    
    def validate_mapping(self, mapping: Dict[str, str]) -> Dict[str, Any]:
        """
        验证字段映射配置
        
        Args:
            mapping: 字段映射 {"excel_column": "system_field", ...}
        
        Returns:
            {
                "valid": True/False,
                "missing_required": ["field1", ...],
                "errors": [...]
            }
        """
        errors = []
        missing_required = []
        
        # 获取已映射的系统字段
        mapped_fields = set(mapping.values())
        
        # 检查必填字段是否都已映射
        fields_config = self.field_config.get("fields", {})
        for field_name, field_info in fields_config.items():
            if field_info.get("required", False):
                if field_name not in mapped_fields:
                    missing_required.append(field_info.get("label", field_name))
        
        if missing_required:
            errors.append(f"缺少必填字段: {', '.join(missing_required)}")
        
        return {
            "valid": len(errors) == 0,
            "missing_required": missing_required,
            "errors": errors
        }
    
    def import_data(
        self,
        file_content: bytes,
        mapping: Dict[str, str],
        db: Session,
        model_class: Any,
        validate_func: Optional[callable] = None,
        progress_callback: Optional[callable] = None
    ) -> Dict[str, Any]:
        """
        执行数据导入
        
        Args:
            file_content: Excel文件内容
            mapping: 字段映射
            db: 数据库会话
            model_class: 数据模型类
            validate_func: 自定义验证函数
        
        Returns:
            {
                "success_count": 80,
                "failed_count": 20,
                "failed_items": [...]
            }
        """
        # 验证文件内容
        if not file_content or len(file_content) == 0:
            raise ValueError("文件内容为空")
        
        # 读取Excel
        try:
            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            error_msg = str(e)
            if "zip" in error_msg.lower():
                raise ValueError("文件格式错误，请上传有效的 Excel 文件（.xlsx 格式）")
            else:
                raise ValueError(f"无法读取 Excel 文件: {error_msg}")
        
        if not rows:
            raise ValueError("Excel文件为空")
        
        headers = [str(cell) if cell is not None else f"列{i+1}" 
                  for i, cell in enumerate(rows[0])]
        data_rows = rows[1:]
        
        success_count = 0
        failed_count = 0
        failed_items = []
        total_rows = len(data_rows)
        
        # 逐行导入
        for row_idx, row in enumerate(data_rows, start=2):  # 从第2行开始（第1行是表头）
            # 报告进度
            if progress_callback:
                progress_callback(row_idx - 1, total_rows)
            try:
                # 构建数据字典
                row_data = {}
                for excel_col, system_field in mapping.items():
                    if system_field:  # 跳过未映射的列
                        col_idx = headers.index(excel_col) if excel_col in headers else -1
                        if col_idx >= 0 and col_idx < len(row):
                            value = row[col_idx]
                            row_data[system_field] = str(value) if value is not None else ""
                
                # 验证数据
                validation_error = self._validate_row_data(row_data)
                if validation_error:
                    failed_count += 1
                    failed_items.append({
                        "row": row_idx,
                        "data": row_data,
                        "reason": validation_error
                    })
                    continue
                
                # 自定义验证
                if validate_func:
                    custom_error = validate_func(row_data, db)
                    if custom_error:
                        failed_count += 1
                        failed_items.append({
                            "row": row_idx,
                            "data": row_data,
                            "reason": custom_error
                        })
                        continue
                
                # 创建模型实例
                instance = model_class(**row_data)
                db.add(instance)
                db.flush()  # 立即执行以检查唯一性约束
                
                success_count += 1
                
            except Exception as e:
                failed_count += 1
                failed_items.append({
                    "row": row_idx,
                    "data": row_data if 'row_data' in locals() else {},
                    "reason": str(e)
                })
                db.rollback()  # 回滚当前行的操作
        
        # 提交成功的记录
        if success_count > 0:
            db.commit()
        
        return {
            "success_count": success_count,
            "failed_count": failed_count,
            "failed_items": failed_items[:100]  # 最多返回100条失败记录
        }
    
    def _validate_row_data(self, row_data: Dict[str, Any]) -> Optional[str]:
        """
        验证行数据
        
        Args:
            row_data: 行数据
        
        Returns:
            错误信息，如果验证通过则返回None
        """
        fields_config = self.field_config.get("fields", {})
        
        for field_name, field_info in fields_config.items():
            value = row_data.get(field_name, "")
            
            # 必填检查
            if field_info.get("required", False):
                if not value or str(value).strip() == "":
                    return f"{field_info.get('label', field_name)}不能为空"
            
            # 长度检查
            max_length = field_info.get("max_length")
            if max_length and len(str(value)) > max_length:
                return f"{field_info.get('label', field_name)}长度不能超过{max_length}"
        
        return None

    def import_data_with_progress(
        self,
        file_content: bytes,
        mapping: Dict[str, str],
        db: Session,
        model_class: Any,
        validate_func: Optional[callable] = None,
        progress_callback: Optional[callable] = None
    ) -> Dict[str, Any]:
        """
        执行数据导入（带进度回调）
        
        这是 import_data 的别名，为了更清晰的语义
        """
        return self.import_data(
            file_content,
            mapping,
            db,
            model_class,
            validate_func,
            progress_callback
        )
