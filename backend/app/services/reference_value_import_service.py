"""
参考价值智能导入服务
"""
from typing import Dict, List, Any, Optional
from io import BytesIO
from decimal import Decimal, InvalidOperation
import openpyxl
from sqlalchemy.orm import Session
from difflib import SequenceMatcher
import uuid

from app.models.department import Department
from app.models.reference_value import ReferenceValue


class ReferenceValueImportService:
    """参考价值智能导入服务"""
    
    # 会话存储（生产环境应使用Redis）
    _sessions: Dict[str, Dict[str, Any]] = {}
    
    @classmethod
    def parse_excel(cls, file_content: bytes, sheet_name: Optional[str] = None, skip_rows: int = 0) -> Dict[str, Any]:
        """
        第一步：解析Excel文件，返回列名和预览数据
        """
        if not file_content or len(file_content) == 0:
            raise ValueError("文件内容为空")
        
        try:
            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        except Exception as e:
            error_msg = str(e)
            if "zip" in error_msg.lower():
                raise ValueError("文件格式错误，请上传有效的 Excel 文件（.xlsx 格式）")
            else:
                raise ValueError(f"无法读取 Excel 文件: {error_msg}")
        
        sheet_names = wb.sheetnames
        
        if sheet_name and sheet_name in sheet_names:
            ws = wb[sheet_name]
            current_sheet = sheet_name
        else:
            ws = wb.active
            current_sheet = ws.title
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel文件为空")
        
        if skip_rows > 0:
            if skip_rows >= len(rows):
                raise ValueError(f"跳过行数({skip_rows})超过总行数({len(rows)})")
            rows = rows[skip_rows:]
        
        if not rows:
            raise ValueError("跳过指定行数后没有剩余数据")
        
        def get_excel_column_name(col_index):
            result = ""
            while col_index >= 0:
                result = chr(col_index % 26 + 65) + result
                col_index = col_index // 26 - 1
            return result
        
        headers = []
        for i, cell in enumerate(rows[0]):
            col_name = get_excel_column_name(i)
            if cell is not None and str(cell).strip():
                headers.append(f"{str(cell).strip()} ({col_name})")
            else:
                headers.append(f"(空列-{col_name})")
        
        data_rows = rows[1:]
        total_rows = len(data_rows)
        
        if total_rows == 0:
            raise ValueError("Excel文件没有数据行")
        
        preview_data = []
        for row in data_rows[:10]:
            preview_data.append([str(cell).strip() if cell is not None else "" for cell in row])
        
        suggested_mapping = cls._suggest_field_mapping(headers)
        
        session_id = str(uuid.uuid4())
        cls._sessions[session_id] = {
            "file_content": file_content,
            "sheet_name": current_sheet,
            "skip_rows": skip_rows,
            "headers": headers,
            "total_rows": total_rows
        }
        
        return {
            "session_id": session_id,
            "sheet_names": sheet_names,
            "current_sheet": current_sheet,
            "headers": headers,
            "preview_data": preview_data,
            "total_rows": total_rows,
            "skip_rows": skip_rows,
            "suggested_mapping": suggested_mapping
        }
    
    @classmethod
    def _suggest_field_mapping(cls, headers: List[str]) -> Dict[str, str]:
        """根据表头建议字段映射"""
        suggested = {}
        
        keywords_map = {
            "period": ["年月", "月份", "期间", "日期", "period"],
            "department_code": ["科室代码", "科室编码", "部门代码", "部门编码", "dept_code"],
            "department_name": ["科室名称", "科室", "部门名称", "部门", "dept_name"],
            "reference_value": ["参考价值", "参考总价值", "总价值", "价值", "reference"],
            "doctor_reference_value": ["医生参考", "医生价值", "医生", "doctor"],
            "nurse_reference_value": ["护理参考", "护理价值", "护理", "nurse"],
            "tech_reference_value": ["医技参考", "医技价值", "医技", "tech"]
        }
        
        for header in headers:
            header_lower = header.lower()
            for field, keywords in keywords_map.items():
                if field in suggested:
                    continue
                for keyword in keywords:
                    if keyword.lower() in header_lower:
                        suggested[field] = header
                        break
        
        return suggested
    
    @classmethod
    def extract_unique_values(
        cls,
        session_id: str,
        field_mapping: Dict[str, str],
        db: Session,
        hospital_id: int,
        match_by: str = "code"
    ) -> Dict[str, Any]:
        """
        第二步：提取科室名称的唯一值，并提供智能匹配建议（仅在按名称匹配时使用）
        """
        session_data = cls._sessions.get(session_id)
        if not session_data:
            raise ValueError("会话已过期或不存在")
        
        # 验证必需字段
        required_fields = ["department_code" if match_by == "code" else "department_name", "reference_value"]
        for field in required_fields:
            if field not in field_mapping:
                raise ValueError(f"必须映射{field}字段")
        
        file_content = session_data["file_content"]
        headers = session_data["headers"]
        sheet_name = session_data.get("sheet_name")
        skip_rows = session_data.get("skip_rows", 0)
        
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if skip_rows > 0:
            rows = rows[skip_rows:]
        data_rows = rows[1:]
        
        # 获取系统科室列表
        system_departments = cls._get_system_departments(db, hospital_id)
        
        # 保存到会话
        cls._sessions[session_id]["field_mapping"] = field_mapping
        cls._sessions[session_id]["match_by"] = match_by
        
        # 如果是按代码匹配，直接返回空的唯一值列表（跳过第二步）
        if match_by == "code":
            return {
                "unique_values": [],
                "system_departments": system_departments
            }
        
        # 按名称匹配时，提取唯一的科室名称
        dept_name_col = field_mapping.get("department_name")
        if not dept_name_col or dept_name_col not in headers:
            raise ValueError("必须映射科室名称字段")
        
        col_idx = headers.index(dept_name_col)
        
        unique_values_dict = {}
        for row in data_rows:
            if col_idx < len(row) and row[col_idx]:
                value = str(row[col_idx]).strip()
                if value:
                    if value not in unique_values_dict:
                        unique_values_dict[value] = 0
                    unique_values_dict[value] += 1
        
        # 为每个唯一值提供智能匹配建议
        unique_values = []
        for value, count in unique_values_dict.items():
            suggested_depts = cls._suggest_departments(value, system_departments)
            unique_values.append({
                "value": value,
                "count": count,
                "suggested_departments": suggested_depts
            })
        
        unique_values.sort(key=lambda x: -x["count"])
        
        cls._sessions[session_id]["unique_values"] = unique_values
        
        return {
            "unique_values": unique_values,
            "system_departments": system_departments
        }
    
    @classmethod
    def _get_system_departments(cls, db: Session, hospital_id: int) -> List[Dict[str, Any]]:
        """获取系统科室列表（使用核算单元代码和名称）"""
        departments = db.query(Department).filter(
            Department.hospital_id == hospital_id
        ).all()
        
        result = []
        for dept in departments:
            # 优先使用核算单元代码和名称，如果没有则使用HIS代码和名称
            code = dept.accounting_unit_code or dept.his_code
            name = dept.accounting_unit_name or dept.his_name
            if code and name:  # 确保代码和名称都存在
                result.append({
                    "id": dept.id,
                    "code": code,
                    "name": name,
                    "his_code": dept.his_code,  # 保留HIS代码用于存储
                    "his_name": dept.his_name,  # 保留HIS名称用于存储
                    "score": 1.0
                })
        
        return result
    
    @classmethod
    def _suggest_departments(
        cls,
        value: str,
        system_departments: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """为给定值建议匹配的科室"""
        suggestions = []
        value_lower = value.lower()
        
        for dept in system_departments:
            name_similarity = SequenceMatcher(None, value_lower, dept["name"].lower()).ratio()
            code_similarity = SequenceMatcher(None, value_lower, dept["code"].lower()).ratio()
            
            contains_in_name = value_lower in dept["name"].lower() or dept["name"].lower() in value_lower
            
            score = max(name_similarity, code_similarity)
            if contains_in_name:
                score += 0.3
            
            # 完全匹配
            if value == dept["name"] or value == dept["code"]:
                score = 1.0
            
            if score > 0.3:
                suggestions.append({
                    "id": dept["id"],
                    "code": dept["code"],
                    "name": dept["name"],
                    "score": score
                })
        
        suggestions.sort(key=lambda x: x["score"], reverse=True)
        return suggestions[:5]
    
    @classmethod
    def generate_preview(
        cls,
        session_id: str,
        value_mapping: List[Dict[str, Any]],
        db: Session,
        hospital_id: int
    ) -> Dict[str, Any]:
        """
        第三步：生成导入预览
        """
        session_data = cls._sessions.get(session_id)
        if not session_data:
            raise ValueError("会话已过期或不存在")
        
        field_mapping = session_data.get("field_mapping")
        if not field_mapping:
            raise ValueError("缺少字段映射信息")
        
        match_by = session_data.get("match_by", "code")
        
        # 构建值到科室代码的映射字典（仅用于名称匹配）
        value_to_dept = {}
        for mapping in value_mapping:
            if hasattr(mapping, 'value'):
                value_to_dept[mapping.value] = mapping.department_code
            else:
                value_to_dept[mapping["value"]] = mapping.get("department_code")
        
        file_content = session_data["file_content"]
        headers = session_data["headers"]
        sheet_name = session_data.get("sheet_name")
        skip_rows = session_data.get("skip_rows", 0)
        
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if skip_rows > 0:
            rows = rows[skip_rows:]
        data_rows = rows[1:]
        
        # 获取列索引
        def get_col_idx(field_name):
            col_name = field_mapping.get(field_name)
            if col_name and col_name in headers:
                return headers.index(col_name)
            return -1
        
        period_col = get_col_idx("period")
        dept_code_col = get_col_idx("department_code")
        dept_name_col = get_col_idx("department_name")
        ref_value_col = get_col_idx("reference_value")
        doctor_col = get_col_idx("doctor_reference_value")
        nurse_col = get_col_idx("nurse_reference_value")
        tech_col = get_col_idx("tech_reference_value")
        
        # 获取系统科室（使用核算单元代码和名称进行匹配）
        all_departments = db.query(Department).filter(Department.hospital_id == hospital_id).all()
        
        # 按核算单元代码索引（优先）和HIS代码索引（备用）
        all_departments_by_code = {}
        all_departments_by_name = {}
        for dept in all_departments:
            # 核算单元代码优先
            if dept.accounting_unit_code:
                all_departments_by_code[dept.accounting_unit_code] = dept
            # HIS代码作为备用
            if dept.his_code:
                all_departments_by_code.setdefault(dept.his_code, dept)
            # 核算单元名称优先
            if dept.accounting_unit_name:
                all_departments_by_name[dept.accounting_unit_name] = dept
            # HIS名称作为备用
            if dept.his_name:
                all_departments_by_name.setdefault(dept.his_name, dept)
        
        # 获取已存在的参考价值
        existing_refs = {}
        for ref in db.query(ReferenceValue).filter(ReferenceValue.hospital_id == hospital_id).all():
            existing_refs[(ref.period, ref.department_code)] = ref
        
        preview_items = []
        statistics = {"total": 0, "new_count": 0, "update_count": 0, "error_count": 0}
        
        def get_cell_value(row, col_idx):
            if col_idx >= 0 and col_idx < len(row) and row[col_idx] is not None:
                return str(row[col_idx]).strip()
            return ""
        
        def parse_decimal(value_str):
            if not value_str:
                return None
            try:
                return Decimal(value_str.replace(",", ""))
            except (InvalidOperation, ValueError):
                return None
        
        for row in data_rows:
            # 获取年月
            period = get_cell_value(row, period_col)
            if not period:
                continue
            
            # 标准化年月格式
            period = cls._normalize_period(period)
            if not period:
                continue
            
            # 获取科室信息
            excel_dept_name = ""
            department = None
            
            if match_by == "code":
                dept_code = get_cell_value(row, dept_code_col)
                if not dept_code:
                    continue
                department = all_departments_by_code.get(dept_code)
                excel_dept_name = get_cell_value(row, dept_name_col) if dept_name_col >= 0 else dept_code
            else:
                excel_dept_name = get_cell_value(row, dept_name_col)
                if not excel_dept_name:
                    continue
                # 从映射中获取科室代码
                dept_code = value_to_dept.get(excel_dept_name)
                if dept_code:
                    department = all_departments_by_code.get(dept_code)
            
            # 获取参考价值
            ref_value = parse_decimal(get_cell_value(row, ref_value_col))
            if ref_value is None:
                continue
            
            doctor_value = parse_decimal(get_cell_value(row, doctor_col))
            nurse_value = parse_decimal(get_cell_value(row, nurse_col))
            tech_value = parse_decimal(get_cell_value(row, tech_col))
            
            statistics["total"] += 1
            
            # 确定状态
            if not department:
                status = "error"
                message = f"科室不存在：{excel_dept_name}"
                dept_code = ""
                dept_name = excel_dept_name
                statistics["error_count"] += 1
            else:
                # 使用核算单元代码和名称（如果有），否则使用HIS代码和名称
                dept_code = department.accounting_unit_code or department.his_code
                dept_name = department.accounting_unit_name or department.his_name
                
                if (period, dept_code) in existing_refs:
                    status = "update"
                    message = "将覆盖已有数据"
                    statistics["update_count"] += 1
                else:
                    status = "new"
                    message = ""
                    statistics["new_count"] += 1
            
            preview_items.append({
                "period": period,
                "department_code": dept_code,
                "department_name": dept_name,
                "excel_department_name": excel_dept_name,
                "reference_value": ref_value,
                "doctor_reference_value": doctor_value,
                "nurse_reference_value": nurse_value,
                "tech_reference_value": tech_value,
                "status": status,
                "message": message
            })
        
        # 按 (period, department_code) 合并，数值累加
        merged_items = {}
        for item in preview_items:
            key = (item["period"], item["department_code"])
            if key in merged_items:
                # 累加数值
                existing = merged_items[key]
                existing["reference_value"] = (existing["reference_value"] or Decimal(0)) + (item["reference_value"] or Decimal(0))
                if item["doctor_reference_value"] is not None:
                    existing["doctor_reference_value"] = (existing["doctor_reference_value"] or Decimal(0)) + item["doctor_reference_value"]
                if item["nurse_reference_value"] is not None:
                    existing["nurse_reference_value"] = (existing["nurse_reference_value"] or Decimal(0)) + item["nurse_reference_value"]
                if item["tech_reference_value"] is not None:
                    existing["tech_reference_value"] = (existing["tech_reference_value"] or Decimal(0)) + item["tech_reference_value"]
                # 更新消息提示合并了多少条
                merge_count = existing.get("_merge_count", 1) + 1
                existing["_merge_count"] = merge_count
                existing["message"] = f"合并了 {merge_count} 条记录" + (", 将覆盖已有数据" if existing["status"] == "update" else "")
            else:
                merged_items[key] = item.copy()
                merged_items[key]["_merge_count"] = 1
        
        # 移除内部计数字段，重新计算统计
        final_items = []
        final_statistics = {"total": 0, "new_count": 0, "update_count": 0, "error_count": 0}
        for item in merged_items.values():
            item.pop("_merge_count", None)
            final_items.append(item)
            final_statistics["total"] += 1
            if item["status"] == "new":
                final_statistics["new_count"] += 1
            elif item["status"] == "update":
                final_statistics["update_count"] += 1
            elif item["status"] == "error":
                final_statistics["error_count"] += 1
        
        cls._sessions[session_id]["preview_items"] = final_items
        
        return {
            "preview_items": final_items,
            "statistics": final_statistics
        }
    
    @classmethod
    def _normalize_period(cls, period_str: str) -> Optional[str]:
        """标准化年月格式为 YYYY-MM"""
        import re
        
        # 尝试匹配各种格式
        # YYYY-MM
        if re.match(r'^\d{4}-\d{2}$', period_str):
            return period_str
        
        # YYYY/MM
        match = re.match(r'^(\d{4})/(\d{1,2})$', period_str)
        if match:
            return f"{match.group(1)}-{match.group(2).zfill(2)}"
        
        # YYYYMM
        match = re.match(r'^(\d{4})(\d{2})$', period_str)
        if match:
            return f"{match.group(1)}-{match.group(2)}"
        
        # YYYY年MM月
        match = re.match(r'^(\d{4})年(\d{1,2})月?$', period_str)
        if match:
            return f"{match.group(1)}-{match.group(2).zfill(2)}"
        
        return None
    
    @classmethod
    def execute_import(
        cls,
        session_id: str,
        confirmed_items: Optional[List[Dict[str, Any]]],
        db: Session,
        hospital_id: int
    ) -> Dict[str, Any]:
        """执行导入"""
        import logging
        logger = logging.getLogger(__name__)
        
        if confirmed_items:
            items_to_import = confirmed_items
        else:
            session_data = cls._sessions.get(session_id)
            if not session_data:
                raise ValueError("会话已过期或不存在")
            
            preview_items = session_data.get("preview_items")
            if not preview_items:
                raise ValueError("缺少预览数据")
            
            items_to_import = preview_items
        
        success_count = 0
        update_count = 0
        error_count = 0
        errors = []
        
        def get_field(item, field_name, default=None):
            if hasattr(item, field_name):
                return getattr(item, field_name, default)
            elif isinstance(item, dict):
                return item.get(field_name, default)
            return default
        
        for item in items_to_import:
            try:
                status = get_field(item, "status")
                if status == "error":
                    error_count += 1
                    errors.append({
                        "period": get_field(item, "period", ""),
                        "department_code": get_field(item, "department_code", ""),
                        "reason": get_field(item, "message", "未知错误")
                    })
                    continue
                
                period = get_field(item, "period")
                dept_code = get_field(item, "department_code")
                dept_name = get_field(item, "department_name")
                ref_value = get_field(item, "reference_value")
                doctor_value = get_field(item, "doctor_reference_value")
                nurse_value = get_field(item, "nurse_reference_value")
                tech_value = get_field(item, "tech_reference_value")
                
                # 检查是否已存在
                existing = db.query(ReferenceValue).filter(
                    ReferenceValue.hospital_id == hospital_id,
                    ReferenceValue.period == period,
                    ReferenceValue.department_code == dept_code
                ).first()
                
                if existing:
                    # 更新
                    existing.department_name = dept_name
                    existing.reference_value = ref_value
                    existing.doctor_reference_value = doctor_value
                    existing.nurse_reference_value = nurse_value
                    existing.tech_reference_value = tech_value
                    update_count += 1
                else:
                    # 新增
                    new_ref = ReferenceValue(
                        hospital_id=hospital_id,
                        period=period,
                        department_code=dept_code,
                        department_name=dept_name,
                        reference_value=ref_value,
                        doctor_reference_value=doctor_value,
                        nurse_reference_value=nurse_value,
                        tech_reference_value=tech_value
                    )
                    db.add(new_ref)
                    success_count += 1
                
                db.flush()
                
            except Exception as e:
                db.rollback()
                error_count += 1
                errors.append({
                    "period": get_field(item, "period", ""),
                    "department_code": get_field(item, "department_code", ""),
                    "reason": str(e)
                })
        
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            raise ValueError(f"提交事务失败: {str(e)}")
        
        if session_id and session_id in cls._sessions:
            del cls._sessions[session_id]
        
        return {
            "success": True,
            "report": {
                "success_count": success_count,
                "update_count": update_count,
                "error_count": error_count,
                "errors": errors[:100]
            }
        }
