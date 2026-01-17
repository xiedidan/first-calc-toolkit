"""
导入人员经费数据到cost_values表和cost_reports表

数据来源：data/25年基本工资/*.xlsx
人员经费 = E列 + G列 + H列 + I列 + J列
科室匹配：B列（去除前导空格）与硬编码的科室对照表匹配
"""
import os
import re
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text
from dotenv import load_dotenv

# 配置
HOSPITAL_ID = 1
DATA_DIR = Path("data/25年基本工资")

# 人员经费维度（医生、护理、医技三个序列）
HR_DIMENSIONS = [
    ("dim-doc-cost-hr", "人员经费"),
    ("dim-nur-cost-hr", "人员经费"),
    ("dim-tech-cost-hr", "人员经费"),
]

# 科室对照表：Excel中的科室名称 -> (核算单元代码, 核算单元名称)
DEPT_MAPPING = {
    "白内障专科": ("YS01", "白内障专科"),
    "青光眼专科": ("YS03", "青光眼专科"),
    "角膜病专科": ("YS02", "角膜病专科"),
    "屈光手术专科": ("YS04", "屈光手术专科"),
    "斜视与小儿眼病专科": ("YS05", "斜视与小儿眼病专科"),
    "眼底病外科专科": ("YS07", "眼底病外科"),
    "眼底病内科专科": ("YS06", "眼底病内科"),
    "眼视光专科": ("YS08", "眼视光专科"),
    "眼外伤专科": ("YS09", "眼外伤专科"),
    "眼整形眼眶病专科": ("YS10", "眼整形眼眶病专科"),
    "中医眼科": ("DUM-01", "中医科"),
    "内科": ("YS13", "内科门诊"),
    "化验室": ("YJ03", "化验室"),
    "放射科": ("YJ01", "放射"),
    "功检科": ("YJ02", "功检室"),
    "手术室": ("FHL01", "手术窒"),
    "消毒供应室": ("FHL06", "供应室"),
    "一病区": ("BHL01", "眼科一病区"),
    "二病区": ("BHL02", "眼科二病区"),
    "日间病区": ("BHL04", "日间病区"),
    "日间玻注中心": ("BHL05", "日间病区(玻注)"),
    "治疗室": ("FHL03", "治疗室"),
    "西药房": ("YJ06", "西药房"),
    "中药房": ("YJ07", "中药房"),
    "医学验光": ("YXYG01", "医学验光"),
    "干眼门诊": ("FHL07", "干眼门诊"),
    "眼底激光室": ("FHL08", "眼底激光室"),
    "荧光造影室": ("FHL09", "荧光造影室"),
    "检验科": ("YJ03", "化验室"),
    "医学影像科": ("YJ01", "放射"),
    "药剂科": ("YJ06", "西药房"),
    "视觉康复中心": ("YXYG01", "医学验光"),
}

def get_year_month_from_filename(filename: str) -> str:
    """从文件名提取年月，转换为 YYYY-MM 格式"""
    # 文件名格式: 25.01.xlsx -> 2025-01
    match = re.match(r"(\d{2})\.(\d{2})\.xlsx", filename)
    if match:
        year = int(match.group(1))
        month = match.group(2)
        full_year = 2000 + year if year < 50 else 1900 + year
        return f"{full_year}-{month}"
    return None

def parse_excel_file(filepath: Path) -> dict:
    """解析Excel文件，返回科室人员经费数据字典（同名科室合并）"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # 使用字典合并同名科室
    dept_costs = {}
    
    # 数据从第6行开始（前5行是标题）
    for row_num in range(6, ws.max_row + 1):
        dept_name = ws.cell(row_num, 2).value  # B列：单位名称
        if not dept_name:
            continue
        
        # 去除前导空格（全角和半角）
        dept_name = dept_name.strip().replace('\u3000', '').strip()
        if not dept_name or dept_name in ['合计', '小计', '总计']:
            continue
        
        # E、G、H、I、J列的值
        col_e = ws.cell(row_num, 5).value or 0  # 岗位工资
        col_g = ws.cell(row_num, 7).value or 0  # 薪级工资
        col_h = ws.cell(row_num, 8).value or 0  # 工资10%
        col_i = ws.cell(row_num, 9).value or 0  # 基本工资
        col_j = ws.cell(row_num, 10).value or 0  # 岗级工资
        
        # 转换为数值
        def to_float(v):
            if v is None:
                return 0.0
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).replace(',', ''))
            except:
                return 0.0
        
        hr_cost = to_float(col_e) + to_float(col_g) + to_float(col_h) + to_float(col_i) + to_float(col_j)
        
        if hr_cost > 0:
            # 合并同名科室
            if dept_name in dept_costs:
                dept_costs[dept_name] += hr_cost
            else:
                dept_costs[dept_name] = hr_cost
    
    wb.close()
    return dept_costs

def main():
    load_dotenv("backend/.env")
    engine = create_engine(os.getenv("DATABASE_URL"))
    
    # 获取所有Excel文件（排除临时文件）
    excel_files = sorted([
        f for f in DATA_DIR.glob("*.xlsx") 
        if not f.name.startswith("~$")
    ])
    
    print(f"找到 {len(excel_files)} 个Excel文件")
    print("=" * 60)
    
    with engine.connect() as conn:
        print(f"使用硬编码的 {len(DEPT_MAPPING)} 个科室映射")
        print()
        
        # cost_values 统计
        cv_total_inserted = 0
        cv_total_updated = 0
        # cost_reports 统计
        cr_total_inserted = 0
        cr_total_updated = 0
        unmatched_depts = set()
        
        for filepath in excel_files:
            year_month = get_year_month_from_filename(filepath.name)
            if not year_month:
                print(f"⚠ 跳过文件（无法解析年月）: {filepath.name}")
                continue
            
            print(f"处理文件: {filepath.name} -> {year_month}")
            dept_costs = parse_excel_file(filepath)
            print(f"  解析到 {len(dept_costs)} 个科室（已合并同名）")
            
            cv_file_inserted = 0
            cv_file_updated = 0
            cr_file_inserted = 0
            cr_file_updated = 0
            
            for dept_name, cost_value in dept_costs.items():
                # 匹配核算单元
                if dept_name not in DEPT_MAPPING:
                    unmatched_depts.add(dept_name)
                    continue
                
                dept_code, dept_display_name = DEPT_MAPPING[dept_name]
                cost_value = round(cost_value, 2)
                
                # ========== 1. 导入 cost_values 表 ==========
                for dim_code, dim_name in HR_DIMENSIONS:
                    existing = conn.execute(text("""
                        SELECT id FROM cost_values 
                        WHERE hospital_id = :hospital_id 
                          AND year_month = :year_month
                          AND dept_code = :dept_code
                          AND dimension_code = :dimension_code
                    """), {
                        "hospital_id": HOSPITAL_ID,
                        "year_month": year_month,
                        "dept_code": dept_code,
                        "dimension_code": dim_code
                    }).fetchone()
                    
                    if existing:
                        conn.execute(text("""
                            UPDATE cost_values 
                            SET cost_value = :cost_value,
                                dept_name = :dept_name,
                                updated_at = NOW()
                            WHERE id = :id
                        """), {
                            "id": existing[0],
                            "cost_value": cost_value,
                            "dept_name": dept_display_name
                        })
                        cv_file_updated += 1
                    else:
                        conn.execute(text("""
                            INSERT INTO cost_values 
                            (hospital_id, year_month, dept_code, dept_name, 
                             dimension_code, dimension_name, cost_value, created_at, updated_at)
                            VALUES 
                            (:hospital_id, :year_month, :dept_code, :dept_name,
                             :dimension_code, :dimension_name, :cost_value, NOW(), NOW())
                        """), {
                            "hospital_id": HOSPITAL_ID,
                            "year_month": year_month,
                            "dept_code": dept_code,
                            "dept_name": dept_display_name,
                            "dimension_code": dim_code,
                            "dimension_name": dim_name,
                            "cost_value": cost_value
                        })
                        cv_file_inserted += 1
                
                # ========== 2. 导入 cost_reports 表 ==========
                existing_cr = conn.execute(text("""
                    SELECT id FROM cost_reports 
                    WHERE hospital_id = :hospital_id 
                      AND period = :period
                      AND department_code = :department_code
                """), {
                    "hospital_id": HOSPITAL_ID,
                    "period": year_month,
                    "department_code": dept_code
                }).fetchone()
                
                if existing_cr:
                    conn.execute(text("""
                        UPDATE cost_reports 
                        SET personnel_cost = :personnel_cost,
                            department_name = :department_name,
                            updated_at = NOW()
                        WHERE id = :id
                    """), {
                        "id": existing_cr[0],
                        "personnel_cost": cost_value,
                        "department_name": dept_display_name
                    })
                    cr_file_updated += 1
                else:
                    conn.execute(text("""
                        INSERT INTO cost_reports 
                        (hospital_id, period, department_code, department_name, 
                         personnel_cost, material_cost, medicine_cost, depreciation_cost, other_cost,
                         created_at, updated_at)
                        VALUES 
                        (:hospital_id, :period, :department_code, :department_name,
                         :personnel_cost, 0, 0, 0, 0, NOW(), NOW())
                    """), {
                        "hospital_id": HOSPITAL_ID,
                        "period": year_month,
                        "department_code": dept_code,
                        "department_name": dept_display_name,
                        "personnel_cost": cost_value
                    })
                    cr_file_inserted += 1
            
            conn.commit()
            print(f"  cost_values: 插入 {cv_file_inserted}, 更新 {cv_file_updated}")
            print(f"  cost_reports: 插入 {cr_file_inserted}, 更新 {cr_file_updated}")
            cv_total_inserted += cv_file_inserted
            cv_total_updated += cv_file_updated
            cr_total_inserted += cr_file_inserted
            cr_total_updated += cr_file_updated
        
        print()
        print("=" * 60)
        print(f"cost_values 总计: 插入 {cv_total_inserted} 条, 更新 {cv_total_updated} 条")
        print(f"cost_reports 总计: 插入 {cr_total_inserted} 条, 更新 {cr_total_updated} 条")
        
        if unmatched_depts:
            print(f"\n⚠ 未匹配的科室 ({len(unmatched_depts)} 个):")
            for dept in sorted(unmatched_depts):
                print(f"  - {dept}")

if __name__ == "__main__":
    main()
