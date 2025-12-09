"""
快速验证导出功能实现
"""
import requests
import openpyxl
from io import BytesIO
from urllib.parse import unquote

BASE_URL = "http://localhost:8000/api/v1"
HOSPITAL_ID = 1

def login():
    response = requests.post(
        f"{BASE_URL}/auth/login",
        json={"username": "admin", "password": "admin123"}
    )
    return response.json()["access_token"] if response.status_code == 200 else None

def get_headers(token):
    return {
        "Authorization": f"Bearer {token}",
        "X-Hospital-ID": str(HOSPITAL_ID),
    }

def verify_export():
    print("验证成本基准导出功能实现")
    print("=" * 60)
    
    token = login()
    if not token:
        print("❌ 登录失败")
        return False
    
    print("✓ 登录成功")
    
    # 创建测试数据
    print("\n创建测试数据...")
    response = requests.get(
        f"{BASE_URL}/model-versions",
        headers=get_headers(token),
        params={"page": 1, "size": 1}
    )
    
    if response.status_code != 200:
        print("❌ 无法获取模型版本")
        return False
    
    version = response.json()["items"][0]
    
    import time
    timestamp = int(time.time())
    
    data = {
        "department_code": f"VERIFY_{timestamp}",
        "department_name": f"验证科室{timestamp}",
        "version_id": version["id"],
        "version_name": version["name"],
        "dimension_code": f"VERIFY_DIM_{timestamp}",
        "dimension_name": f"验证维度{timestamp}",
        "benchmark_value": 1234.56
    }
    
    response = requests.post(
        f"{BASE_URL}/cost-benchmarks",
        headers=get_headers(token),
        json=data
    )
    
    if response.status_code != 200:
        print(f"❌ 创建测试数据失败: {response.text}")
        return False
    
    benchmark_id = response.json()["id"]
    print(f"✓ 创建测试数据成功 (ID: {benchmark_id})")
    
    try:
        # 测试导出
        print("\n测试导出功能...")
        response = requests.get(
            f"{BASE_URL}/cost-benchmarks/export",
            headers=get_headers(token)
        )
        
        if response.status_code != 200:
            print(f"❌ 导出失败: {response.status_code}")
            return False
        
        print("✓ 导出请求成功")
        
        # 验证响应头
        content_type = response.headers.get("Content-Type")
        content_disposition = response.headers.get("Content-Disposition")
        
        print(f"\n响应头验证:")
        print(f"  Content-Type: {content_type}")
        
        if "spreadsheetml.sheet" not in content_type:
            print("  ❌ Content-Type不正确")
            return False
        print("  ✓ Content-Type正确")
        
        decoded_disposition = unquote(content_disposition)
        print(f"  Content-Disposition: {decoded_disposition}")
        
        if "成本基准_" not in decoded_disposition:
            print("  ❌ 文件名不包含中文")
            return False
        print("  ✓ 文件名包含中文")
        
        import re
        if not re.search(r"\d{8}_\d{6}", decoded_disposition):
            print("  ❌ 文件名不包含时间戳")
            return False
        print("  ✓ 文件名包含时间戳")
        
        # 验证Excel内容
        print(f"\nExcel内容验证:")
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # 验证列标题
        expected_headers = [
            "科室代码", "科室名称", "模型版本名称", "维度代码",
            "维度名称", "基准值", "创建时间", "更新时间"
        ]
        actual_headers = [cell.value for cell in ws[1]]
        
        if actual_headers != expected_headers:
            print(f"  ❌ 列标题不正确")
            print(f"    期望: {expected_headers}")
            print(f"    实际: {actual_headers}")
            return False
        print("  ✓ 列标题正确")
        
        # 验证数据
        data_rows = ws.max_row - 1
        print(f"  ✓ 包含{data_rows}行数据")
        
        if data_rows > 0:
            # 查找我们创建的测试数据
            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == f"VERIFY_{timestamp}":
                    found = True
                    print(f"  ✓ 找到测试数据: {row[1]} - {row[4]}")
                    print(f"    基准值: {row[5]}")
                    break
            
            if not found:
                print("  ❌ 未找到测试数据")
                return False
        
        print("\n" + "=" * 60)
        print("✅ 所有验证通过！导出功能实现正确。")
        return True
        
    finally:
        # 清理测试数据
        print(f"\n清理测试数据 (ID: {benchmark_id})...")
        response = requests.delete(
            f"{BASE_URL}/cost-benchmarks/{benchmark_id}",
            headers=get_headers(token)
        )
        if response.status_code == 200:
            print("✓ 清理完成")
        else:
            print(f"⚠ 清理失败: {response.text}")

if __name__ == "__main__":
    success = verify_export()
    exit(0 if success else 1)
